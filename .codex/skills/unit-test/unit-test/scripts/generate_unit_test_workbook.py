#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.cell.cell import MergedCell

_SYNTHETIC_INPUT_SECTION_KEYS = {
    "action",
    "requeststatus",
    "currentstatus",
    "workflowstate",
    "statetransition",
    "requeststate",
}


@dataclass(frozen=True)
class FunctionSheetLayout:
    template_sheet_name: str
    case_start_col: int
    utc_row: int
    condition_row: int
    confirm_row: int
    result_row: int
    passed_cell: str
    failed_cell: str
    untested_cell: str
    n_count_cell: str
    a_count_cell: str
    b_count_cell: str
    total_cell: str


def _read_json_payload(input_path: Optional[str]) -> Dict[str, Any]:
    if input_path:
        text = Path(input_path).read_text(encoding="utf-8")
    else:
        text = sys.stdin.read()

    if not text.strip():
        raise SystemExit("No JSON payload provided (stdin is empty).")

    try:
        payload = json.loads(text)
    except json.JSONDecodeError as exc:
        raise SystemExit(f"Invalid JSON payload: {exc}") from exc

    if not isinstance(payload, dict):
        raise SystemExit("JSON payload must be an object.")

    return payload


def _resolve_default_template_path() -> Path:
    unit_template_dir = Path.cwd() / "docs" / "template" / "unit-test"
    candidates = [
        unit_template_dir / "Report5_Unit Test Case_v2.3.xlsx",
        unit_template_dir / "Report5_Unit Test Case_v1.3.xlsx",
    ]

    for candidate in candidates:
        if candidate.exists():
            return candidate

    if unit_template_dir.exists():
        workbooks = sorted(path for path in unit_template_dir.iterdir() if path.suffix.lower() == ".xlsx")
        if workbooks:
            return workbooks[0]

    return candidates[0]


def _sanitize_sheet_name(name: str) -> str:
    name = (name or "").strip()
    if not name:
        return "Function"

    name = re.sub(r"[:\\/\?\*\[\]]", "_", name)
    if len(name) > 31:
        name = name[:31]
    return name


def _make_unique_sheet_name(existing: Iterable[str], desired: str) -> str:
    existing_set = set(existing)
    if desired not in existing_set:
        return desired

    base = desired[:28] if len(desired) > 28 else desired
    i = 2
    while True:
        candidate = f"{base}-{i}"
        if candidate not in existing_set:
            return candidate
        i += 1


def _find_cell(ws: openpyxl.worksheet.worksheet.Worksheet, needle: str) -> Optional[Tuple[int, int]]:
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value == needle:
                return (cell.row, cell.column)
    return None


def _find_row_col_a(
    ws: openpyxl.worksheet.worksheet.Worksheet, needle: str
) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == needle:
            return row
    return None


def _normalize_input_section_key(value: Any) -> str:
    return re.sub(r"[\s_-]+", "", str(value or "").lower())


def _assert_real_input_name(function_name: str, input_name: str) -> None:
    if _normalize_input_section_key(input_name) not in _SYNTHETIC_INPUT_SECTION_KEYS:
        return

    raise SystemExit(
        f"Function '{function_name or 'Unknown function'}' case input '{input_name}' is a synthetic worksheet section. "
        "Use real request parameters in inputs and move workflow/state notes to preconditions or compact returns."
    )


def _utc_info(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> Optional[Tuple[int, int, int]]:
    pos = _find_cell(ws, "UTCID01")
    if not pos:
        return None
    utc_row, utc_col = pos
    last_col = utc_col
    for col in range(utc_col, ws.max_column + 1):
        v = ws.cell(utc_row, col).value
        if isinstance(v, str) and v.startswith("UTCID"):
            last_col = col
    return utc_row, utc_col, last_col


def _cell_below_label(ws: openpyxl.worksheet.worksheet.Worksheet, label: str) -> str:
    pos = _find_cell(ws, label)
    if not pos:
        raise ValueError(f"Template sheet missing required label: {label}")
    row, col = pos
    return openpyxl.utils.get_column_letter(col) + str(row + 1)


def _detect_layout(wb: openpyxl.Workbook, template_sheet_name: str) -> FunctionSheetLayout:
    ws = wb[template_sheet_name]

    utc = _utc_info(ws)
    if not utc:
        raise SystemExit(f"Template function sheet '{template_sheet_name}' has no UTCID01 row.")
    utc_row, case_start_col, _ = utc

    condition_row = _find_row_col_a(ws, "Condition")
    confirm_row = _find_row_col_a(ws, "Confirm")
    result_row = _find_row_col_a(ws, "Result")
    if not condition_row or not confirm_row or not result_row:
        raise SystemExit(
            f"Template function sheet '{template_sheet_name}' must include 'Condition', 'Confirm', and 'Result' labels in column A."
        )

    # Summary cells are consistently "label row + 1" in the template.
    passed_cell = _cell_below_label(ws, "Passed")
    failed_cell = _cell_below_label(ws, "Failed")
    untested_cell = _cell_below_label(ws, "Untested")
    n_a_b_row = _find_cell(ws, "N/A/B")
    if not n_a_b_row:
        raise SystemExit(f"Template function sheet '{template_sheet_name}' missing 'N/A/B' label.")
    n_count_cell = openpyxl.utils.get_column_letter(n_a_b_row[1]) + str(n_a_b_row[0] + 1)
    a_count_cell = openpyxl.utils.get_column_letter(n_a_b_row[1] + 1) + str(n_a_b_row[0] + 1)
    b_count_cell = openpyxl.utils.get_column_letter(n_a_b_row[1] + 2) + str(n_a_b_row[0] + 1)

    total_pos = _find_cell(ws, "Total Test Cases")
    if not total_pos:
        raise SystemExit(f"Template function sheet '{template_sheet_name}' missing 'Total Test Cases' label.")
    total_cell = openpyxl.utils.get_column_letter(total_pos[1]) + str(total_pos[0] + 1)

    return FunctionSheetLayout(
        template_sheet_name=template_sheet_name,
        case_start_col=case_start_col,
        utc_row=utc_row,
        condition_row=condition_row,
        confirm_row=confirm_row,
        result_row=result_row,
        passed_cell=passed_cell,
        failed_cell=failed_cell,
        untested_cell=untested_cell,
        n_count_cell=n_count_cell,
        a_count_cell=a_count_cell,
        b_count_cell=b_count_cell,
        total_cell=total_cell,
    )


def _pick_template_sheet_name(wb: openpyxl.Workbook, preferred: Optional[str]) -> str:
    if preferred:
        if preferred not in wb.sheetnames:
            raise SystemExit(f"Template sheet not found: {preferred}")
        return preferred

    base_sheets = {"Guideline", "Cover", "Function List", "Test Report"}
    best_name: Optional[str] = None
    best_utc_count = -1
    best_rows = -1

    for name in wb.sheetnames:
        if name in base_sheets:
            continue
        ws = wb[name]
        utc = _utc_info(ws)
        if not utc:
            continue
        if not _find_row_col_a(ws, "Condition") or not _find_row_col_a(ws, "Confirm") or not _find_row_col_a(ws, "Result"):
            continue
        utc_row, utc_col, last_col = utc
        utc_count = 0
        for col in range(utc_col, last_col + 1):
            v = ws.cell(utc_row, col).value
            if isinstance(v, str) and v.startswith("UTCID"):
                utc_count += 1

        if (utc_count, ws.max_row) > (best_utc_count, best_rows):
            best_name = name
            best_utc_count = utc_count
            best_rows = ws.max_row

    if not best_name:
        raise SystemExit("No suitable function sheet template found (needs UTCID01 + Condition/Confirm/Result).")

    return best_name


def _clear_range(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _normalize_scalar(value: Any) -> Any:
    if value is None:
        return "null"
    return value


def _as_list(value: Any) -> List[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def _str_key(value: Any) -> str:
    if isinstance(value, str):
        return value
    try:
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    except TypeError:
        return str(value)


def _write_marks(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col_start: int,
    active_case_indexes: Sequence[int],
) -> None:
    for idx in active_case_indexes:
        ws.cell(row, col_start + idx).value = "O"


def _set_header_value(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    label: str,
    offset_cols: int,
    value: Any,
    search_max_row: int = 10,
    search_max_col: int = 20,
) -> None:
    for row in range(1, min(search_max_row, ws.max_row) + 1):
        for col in range(1, min(search_max_col, ws.max_column) + 1):
            if ws.cell(row, col).value == label:
                ws.cell(row, col + offset_cols).value = value
                return


def _populate_function_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    layout: FunctionSheetLayout,
    fn: Dict[str, Any],
) -> None:
    cases: List[Dict[str, Any]] = fn.get("cases") or []
    if not isinstance(cases, list):
        raise SystemExit(f"Function '{fn.get('function_name')}' cases must be a list.")

    # Overwrite header fields (best-effort, based on known template labels).
    _set_header_value(ws, "Function Code", 2, fn.get("function_code") or "")
    _set_header_value(ws, "Function Name", 6, fn.get("function_name") or "")
    _set_header_value(ws, "Created By", 2, fn.get("created_by") or "")
    _set_header_value(ws, "Executed By", 6, fn.get("executed_by") or (fn.get("created_by") or ""))
    _set_header_value(ws, "Lines  of code", 2, fn.get("loc") or "")
    _set_header_value(ws, "Test requirement", 2, fn.get("test_requirement") or "")

    # Clear matrix area (from UTC row through result block).
    end_row = layout.result_row + 3
    _clear_range(ws, layout.utc_row, end_row, 1, ws.max_column)

    # Recreate static labels.
    ws.cell(layout.condition_row, 1).value = "Condition"
    ws.cell(layout.condition_row, 2).value = "Precondition "
    ws.cell(layout.confirm_row, 1).value = "Confirm"
    ws.cell(layout.confirm_row, 2).value = "Return"
    ws.cell(layout.result_row, 1).value = "Result"
    ws.cell(layout.result_row, 2).value = "Type(N : Normal, A : Abnormal, B : Boundary)"
    ws.cell(layout.result_row + 1, 2).value = "Passed/Failed"
    ws.cell(layout.result_row + 2, 2).value = "Executed Date"
    ws.cell(layout.result_row + 3, 2).value = "Defect ID"

    # UTCID headers.
    for idx in range(len(cases)):
        ws.cell(layout.utc_row, layout.case_start_col + idx).value = f"UTCID{idx + 1:02d}"

    # Build condition unions.
    preconditions_map: Dict[str, List[int]] = {}
    inputs_map: Dict[str, Dict[str, Dict[str, Any]]] = {}
    returns_map: Dict[str, List[int]] = {}
    exceptions_map: Dict[str, List[int]] = {}
    logs_map: Dict[str, List[int]] = {}

    for idx, case in enumerate(cases):
        for pre in _as_list(case.get("preconditions")):
            if not isinstance(pre, str):
                pre = str(pre)
            preconditions_map.setdefault(pre, []).append(idx)

        inputs = case.get("inputs") or {}
        if not isinstance(inputs, dict):
            raise SystemExit(f"Function '{fn.get('function_name')}' case inputs must be an object.")

        for input_name, raw_value in inputs.items():
            _assert_real_input_name(str(fn.get("function_name") or ""), str(input_name))
            if input_name not in inputs_map:
                inputs_map[input_name] = {}
            for v in _as_list(raw_value):
                v = _normalize_scalar(v)
                key = _str_key(v)
                if key not in inputs_map[input_name]:
                    inputs_map[input_name][key] = {"display": v, "cases": []}
                inputs_map[input_name][key]["cases"].append(idx)

        for ret in _as_list(case.get("returns")):
            if not isinstance(ret, str):
                ret = str(ret)
            returns_map.setdefault(ret, []).append(idx)

        for exc in _as_list(case.get("exceptions")):
            if not isinstance(exc, str):
                exc = str(exc)
            exceptions_map.setdefault(exc, []).append(idx)

        for log in _as_list(case.get("logs")):
            if not isinstance(log, str):
                log = str(log)
            logs_map.setdefault(log, []).append(idx)

    # Capacity checks (template has fixed spacing; fail fast if we overflow).
    condition_capacity = layout.confirm_row - layout.condition_row - 1
    condition_required = len(preconditions_map) + sum(1 + len(v) for v in inputs_map.values())
    if condition_required > condition_capacity:
        raise SystemExit(
            f"Function '{fn.get('function_name')}' needs {condition_required} condition rows but template only has {condition_capacity}. "
            "Reduce/merge condition rows or pick a larger template function sheet."
        )

    confirm_capacity = layout.result_row - layout.confirm_row - 1
    confirm_required = (
        len(returns_map)
        + (0 if not exceptions_map else 1 + len(exceptions_map))
        + (0 if not logs_map else 1 + len(logs_map))
    )
    if confirm_required > confirm_capacity:
        raise SystemExit(
            f"Function '{fn.get('function_name')}' needs {confirm_required} confirm rows but template only has {confirm_capacity}. "
            "Reduce/merge return/exception rows or pick a larger template function sheet."
        )

    # Write conditions.
    row = layout.condition_row + 1
    for text, case_indexes in preconditions_map.items():
        ws.cell(row, 4).value = text
        _write_marks(ws, row, layout.case_start_col, case_indexes)
        row += 1

    for input_name, values_map in inputs_map.items():
        ws.cell(row, 2).value = input_name
        row += 1
        for entry in values_map.values():
            ws.cell(row, 4).value = entry["display"]
            _write_marks(ws, row, layout.case_start_col, entry["cases"])
            row += 1

    # Write confirms (Return + Exception + Log message).
    row = layout.confirm_row + 1
    for text, case_indexes in returns_map.items():
        ws.cell(row, 4).value = text
        _write_marks(ws, row, layout.case_start_col, case_indexes)
        row += 1

    if exceptions_map:
        ws.cell(row, 2).value = "Exception"
        row += 1
        for text, case_indexes in exceptions_map.items():
            ws.cell(row, 4).value = text
            _write_marks(ws, row, layout.case_start_col, case_indexes)
            row += 1

    if logs_map:
        ws.cell(row, 2).value = "Log message"
        row += 1
        for text, case_indexes in logs_map.items():
            ws.cell(row, 4).value = text
            _write_marks(ws, row, layout.case_start_col, case_indexes)
            row += 1

    # Write results per case.
    for idx, case in enumerate(cases):
        type_code = (case.get("type") or "").strip().upper()
        status_code = (case.get("status") or "").strip().upper()
        executed_date = case.get("executed_date") or ""
        defect_id = case.get("defect_id") or ""

        ws.cell(layout.result_row, layout.case_start_col + idx).value = type_code
        ws.cell(layout.result_row + 1, layout.case_start_col + idx).value = status_code
        ws.cell(layout.result_row + 2, layout.case_start_col + idx).value = executed_date
        ws.cell(layout.result_row + 3, layout.case_start_col + idx).value = defect_id

    passed_count = sum(1 for case in cases if (case.get("status") or "").strip().upper() == "P")
    failed_count = sum(1 for case in cases if (case.get("status") or "").strip().upper() == "F")
    untested_count = max(len(cases) - passed_count - failed_count, 0)
    n_count = sum(1 for case in cases if (case.get("type") or "").strip().upper() == "N")
    a_count = sum(1 for case in cases if (case.get("type") or "").strip().upper() == "A")
    b_count = sum(1 for case in cases if (case.get("type") or "").strip().upper() == "B")

    ws[layout.passed_cell] = passed_count
    ws[layout.failed_cell] = failed_count
    ws[layout.untested_cell] = untested_count
    ws[layout.n_count_cell] = n_count
    ws[layout.a_count_cell] = a_count
    ws[layout.b_count_cell] = b_count
    ws[layout.total_cell] = len(cases)


def _update_cover(wb: openpyxl.Workbook, meta: Dict[str, Any]) -> None:
    if "Cover" not in wb.sheetnames:
        return
    ws = wb["Cover"]
    if meta.get("project_name"):
        ws["B4"].value = meta["project_name"]
    if meta.get("project_code"):
        ws["B5"].value = meta["project_code"]
    if meta.get("creator"):
        ws["F4"].value = meta["creator"]
    if meta.get("reviewer"):
        ws["F5"].value = meta["reviewer"]
    if meta.get("issue_date"):
        ws["F6"].value = meta["issue_date"]
    if meta.get("version"):
        ws["F7"].value = meta["version"]


def _update_function_list(wb: openpyxl.Workbook, functions: List[Dict[str, Any]], sheet_name_map: Dict[int, str]) -> None:
    if "Function List" not in wb.sheetnames:
        return
    ws = wb["Function List"]

    start_row = 11
    end_row = max(ws.max_row, start_row + len(functions) + 10)
    _clear_range(ws, start_row, end_row, 1, 7)

    for idx, fn in enumerate(functions):
        row = start_row + idx
        ws.cell(row, 1).value = idx + 1
        ws.cell(row, 2).value = fn.get("class_name") or ""
        ws.cell(row, 3).value = fn.get("function_name") or ""
        ws.cell(row, 4).value = fn.get("function_code") or ""
        ws.cell(row, 5).value = sheet_name_map.get(idx, fn.get("sheet_name") or fn.get("function_name") or "")
        ws.cell(row, 6).value = fn.get("description") or ""
        ws.cell(row, 7).value = fn.get("precondition") or ""


def _escape_sheet_ref(sheet_name: str) -> str:
    # Excel formula sheet refs use single quotes; embedded single quote must be doubled.
    return sheet_name.replace("'", "''")


def _update_test_report(wb: openpyxl.Workbook, functions: List[Dict[str, Any]], sheet_name_map: Dict[int, str], layout: FunctionSheetLayout) -> None:
    if "Test Report" not in wb.sheetnames:
        return
    ws = wb["Test Report"]

    start_row = 12
    end_row = max(ws.max_row, start_row + len(functions) + 10)
    _clear_range(ws, start_row, end_row, 1, 9)

    for idx, fn in enumerate(functions):
        row = start_row + idx
        sheet_name = sheet_name_map.get(idx, fn.get("sheet_name") or fn.get("function_name") or "Function")
        sheet_ref = _escape_sheet_ref(sheet_name)

        ws.cell(row, 1).value = idx + 1
        ws.cell(row, 2).value = fn.get("function_code") or ""
        ws.cell(row, 3).value = f"='{sheet_ref}'!{layout.passed_cell}"
        ws.cell(row, 4).value = f"='{sheet_ref}'!{layout.failed_cell}"
        ws.cell(row, 5).value = f"='{sheet_ref}'!{layout.untested_cell}"
        ws.cell(row, 6).value = f"='{sheet_ref}'!{layout.n_count_cell}"
        ws.cell(row, 7).value = f"='{sheet_ref}'!{layout.a_count_cell}"
        ws.cell(row, 8).value = f"='{sheet_ref}'!{layout.b_count_cell}"
        ws.cell(row, 9).value = f"='{sheet_ref}'!{layout.total_cell}"


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Generate a Unit Test Case (UTC) matrix workbook from JSON + template.")
    parser.add_argument("--input", help="Path to JSON payload. If omitted, read from stdin.")
    parser.add_argument(
        "--template",
        default=str(_resolve_default_template_path()),
        help="Path to the XLSX template workbook.",
    )
    parser.add_argument(
        "--template-function-sheet",
        help="Sheet name to use as the per-function sheet template (defaults to the best available sheet).",
    )
    parser.add_argument("--output", required=True, help="Output XLSX path.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite output file if it exists.")
    args = parser.parse_args(argv)

    payload = _read_json_payload(args.input)
    functions = payload.get("functions") or []
    if not isinstance(functions, list):
        raise SystemExit("Payload 'functions' must be a list.")
    if len(functions) != 1:
        raise SystemExit(
            "This generator now expects exactly one functions[] item per workbook. "
            "Merge all UTC cases for the requested business function into a single sheet."
        )

    template_path = Path(args.template)
    if not template_path.exists():
        raise SystemExit(f"Template workbook not found: {template_path}")

    output_path = Path(args.output)
    if output_path.exists() and not args.overwrite:
        raise SystemExit(f"Output already exists: {output_path} (use --overwrite)")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(template_path)
    template_sheet_name = _pick_template_sheet_name(wb, args.template_function_sheet)
    layout = _detect_layout(wb, template_sheet_name)

    for name in list(wb.sheetnames):
        if name != template_sheet_name:
            wb.remove(wb[name])

    template_ws = wb[template_sheet_name]
    template_ws.sheet_state = "visible"

    fn = functions[0]
    if not isinstance(fn, dict):
        raise SystemExit("Each functions[] item must be an object.")
    raw_name = fn.get("sheet_name") or fn.get("function_name") or fn.get("function_code") or "Function"
    desired = _sanitize_sheet_name(str(raw_name))
    actual = _make_unique_sheet_name(
        [name for name in wb.sheetnames if name != template_sheet_name],
        desired,
    )

    if actual != template_sheet_name:
        template_ws.title = actual

    _populate_function_sheet(template_ws, layout, fn)
    wb.defined_names.clear()

    wb.save(output_path)
    print(str(output_path.resolve()))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
