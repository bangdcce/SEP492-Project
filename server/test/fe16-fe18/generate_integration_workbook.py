from __future__ import annotations

import json
from collections import defaultdict
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from case_catalog_data import CASE_CATALOG


REPO_ROOT = Path(__file__).resolve().parents[3]
TEMPLATE_PATH = REPO_ROOT / "docs" / "Integration Test" / "Interation Test.xlsx"
OUTPUT_PATH = REPO_ROOT / "docs" / "Integration Test" / "Integration Test FE16-FE18.xlsx"
EVIDENCE_DIR = REPO_ROOT / "server" / "test-artifacts" / "fe16-fe18" / "evidence"

PROJECT_NAME = "InterDev Platform"
PROJECT_CODE = "INT"
DOC_CODE = "INT_IT_FE16-FE18_v1.0"
ISSUE_DATE = "2026-03-30"
DEFAULT_TESTER = "SonNT"
BLOCKED_CASE_NOTES = {
    "FE16-TPR-01": "Blocked in this environment: Docker Desktop daemon was unavailable, so the Postgres-backed HTTP integration harness could not start.",
    "FE16-TPR-02": "Blocked in this environment: Docker Desktop daemon was unavailable, so the Postgres-backed HTTP integration harness could not start.",
    "FE16-TS-02": "Blocked in this environment: Docker Desktop daemon was unavailable, so the Postgres-backed HTTP integration harness could not start.",
    "FE18-NOT-01": "Blocked in this environment: Docker Desktop daemon was unavailable, so the Postgres-backed HTTP integration harness could not start.",
    "FE18-NOT-02": "Blocked in this environment: Docker Desktop daemon was unavailable, so the Postgres-backed HTTP integration harness could not start.",
    "FE18-NOT-03": "Blocked in this environment: Docker Desktop daemon was unavailable, so the Postgres-backed HTTP integration harness could not start.",
    "FE18-NOT-04": "Blocked in this environment: Docker Desktop daemon was unavailable, so the Postgres-backed HTTP integration harness could not start.",
    "FE18-NOT-05": "Blocked in this environment: Docker Desktop daemon was unavailable, so the Postgres-backed HTTP integration harness could not start.",
    "FE18-NOT-06": "Blocked in this environment: Docker Desktop daemon was unavailable, so the Postgres-backed HTTP integration harness could not start.",
}

FEATURES = [
    {
        "sheet_name": "Rating & Trust Score System",
        "code": "FE-16",
        "requirement": (
            "Validate review creation, rollback behavior, moderation concurrency, "
            "guarded trust profile retrieval, and guarded trust score calculation/history."
        ),
        "references": [
            "server/src/modules/review/review.service.ts",
            "server/src/modules/trust-score/trust-score.controller.ts",
            "server/src/modules/trust-score/trust-score.service.ts",
            "server/src/modules/users/trust-profiles.controller.ts",
            "server/src/modules/users/trust-profiles.service.ts",
            "server/test/fe16-fe18/feature-http.e2e-spec.ts",
        ],
        "sections": [
            "Review Creation & Eligibility",
            "Review Update & Edit History",
            "Review Moderation",
            "Trust Profile Retrieval",
            "Trust Score Calculation & History",
        ],
        "summary": "Reviews, trust profiles, and trust score recalculation/history",
    },
    {
        "sheet_name": "Dispute Resolution System",
        "code": "FE-17",
        "requirement": (
            "Validate dispute intake, evidence validation, hearing lifecycle, verdict issuance, "
            "settlement gates, appeal flows, and staff/admin control signals."
        ),
        "references": [
            "server/src/modules/disputes/disputes.service.ts",
            "server/src/modules/disputes/services/evidence.service.ts",
            "server/src/modules/disputes/services/hearing.service.ts",
            "server/src/modules/disputes/services/hearing-verdict-orchestrator.service.ts",
            "server/src/modules/disputes/services/settlement.service.ts",
            "server/src/modules/disputes/services/staff-assignment.service.ts",
        ],
        "sections": [
            "Dispute Intake & Listing",
            "Evidence Management",
            "Hearing Lifecycle",
            "Verdict Readiness & Issuance",
            "Settlement & Appeal",
            "Staff Assignment / Admin Controls",
        ],
        "summary": "Dispute intake, evidence, hearings, appeals, and admin controls",
    },
    {
        "sheet_name": "Notification System",
        "code": "FE-18",
        "requirement": (
            "Validate guarded notification listing, unread filtering, mark-read APIs, "
            "bulk mark-all-read, and notification persistence plus event dispatch."
        ),
        "references": [
            "server/src/modules/notifications/notifications.controller.ts",
            "server/src/modules/notifications/notifications.service.ts",
            "server/src/modules/disputes/events/notification-realtime.listener.ts",
            "server/test/fe16-fe18/feature-http.e2e-spec.ts",
        ],
        "sections": [
            "Notification Listing & Pagination",
            "Unread Filtering",
            "Mark Single Read",
            "Mark All Read",
            "Notification Creation & Realtime Dispatch",
        ],
        "summary": "Guarded notification APIs and realtime dispatch",
    },
]


def load_catalog() -> list[dict[str, Any]]:
    catalog = [dict(item) for item in CASE_CATALOG]
    evidence_by_id: dict[str, dict[str, Any]] = {}

    if EVIDENCE_DIR.exists():
        for file_path in EVIDENCE_DIR.glob("*.json"):
            evidence_by_id[file_path.stem] = json.loads(file_path.read_text(encoding="utf-8"))

    merged: list[dict[str, Any]] = []
    for item in catalog:
        evidence = evidence_by_id.get(item["id"])
        merged_item = dict(item)
        if evidence:
            merged_item["actualResults"] = evidence.get("actualResults", item.get("actualResults", ""))
            merged_item["result"] = evidence.get("result", item.get("result", "Untested"))
            merged_item["testDate"] = evidence.get("testDate", item.get("testDate", ""))
            merged_item["tester"] = evidence.get("tester", item.get("tester", DEFAULT_TESTER))
            merged_item["note"] = evidence.get("note", item.get("note", ""))
            merged_item["evidenceRef"] = evidence.get("evidenceRef", item.get("evidenceRef", ""))
        elif item["id"] in BLOCKED_CASE_NOTES:
            merged_item["note"] = BLOCKED_CASE_NOTES[item["id"]]
        merged.append(merged_item)
    return merged


def copy_row_style(ws, source_row: int, target_row: int, max_col: int = 26) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
        target.number_format = source.number_format
    if ws.row_dimensions[source_row].height is not None:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def clear_values(ws, start_row: int, end_row: int, max_col: int = 26) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def count_results(cases: list[dict[str, Any]]) -> dict[str, int]:
    counts = {"Pass": 0, "Fail": 0, "Untested": 0, "N/A": 0}
    for case in cases:
        result = case.get("result", "Untested")
        counts[result] = counts.get(result, 0) + 1
    counts["Total"] = len(cases)
    return counts


def detail_actual_text(case: dict[str, Any]) -> str:
    result = case.get("result", "Untested")
    if result == "Untested":
        parts = ["No executed automation evidence was captured for this case in the FE16/17/18 suite."]
        if case.get("note"):
            parts.append(f"Note: {case['note']}")
        return "\n".join(parts)

    parts = [case.get("actualResults", "").strip()]
    parts.append(
        f"Executed on {case.get('testDate') or ISSUE_DATE} by {case.get('tester') or DEFAULT_TESTER}."
    )
    if case.get("evidenceRef"):
        parts.append(f"Evidence: {case['evidenceRef']}.")
    if case.get("note"):
        parts.append(f"Note: {case['note']}")
    return "\n".join(part for part in parts if part)


def populate_cover(ws) -> None:
    ws["C4"] = PROJECT_NAME
    ws["C5"] = PROJECT_CODE
    ws["C6"] = DOC_CODE
    ws["B12"] = ISSUE_DATE
    ws["C12"] = "v1.0"
    ws["D12"] = "Initial FE-16/FE-17/FE-18 integration workbook generated from executed automation evidence."
    ws["E12"] = "A"


def populate_test_case_list(ws, cases: list[dict[str, Any]]) -> None:
    ws["C3"] = PROJECT_NAME
    ws["C4"] = PROJECT_CODE
    ws["C5"] = (
        "1. NestJS backend test harness\n"
        "2. Supertest HTTP client\n"
        "3. Postgres 16 container via testcontainers\n"
        "4. Jest automation suites for review/trust/dispute modules\n"
        "5. Real JWT signing using the same guard configuration"
    )

    clear_values(ws, 9, max(60, ws.max_row), 5)

    grouped = defaultdict(list)
    for case in cases:
        grouped[(case["feature"], case["section"])].append(case)

    row = 9
    sequence = 1
    for feature in FEATURES:
        for section in feature["sections"]:
            section_cases = grouped.get((feature["sheet_name"], section))
            if not section_cases:
                continue
            copy_row_style(ws, 9, row, 5)
            ws.cell(row, 1).value = sequence
            ws.cell(row, 2).value = section
            ws.cell(row, 3).value = feature["sheet_name"]
            ws.cell(row, 4).value = f"{feature['code']} - {len(section_cases)} critical automation-backed cases."
            row += 1
            sequence += 1


def populate_test_report(ws, cases: list[dict[str, Any]]) -> None:
    ws["C3"] = PROJECT_NAME
    ws["C4"] = PROJECT_CODE
    ws["C5"] = DOC_CODE
    ws["C6"] = "Release scope: FE-16 Rating & Trust Score System, FE-17 Dispute Resolution System, FE-18 Notification System."
    ws["H3"] = DEFAULT_TESTER
    ws["H4"] = "N/A"
    ws["H5"] = ISSUE_DATE

    feature_counts = {}
    for feature in FEATURES:
        feature_cases = [case for case in cases if case["feature"] == feature["sheet_name"]]
        feature_counts[feature["sheet_name"]] = count_results(feature_cases)

    for index, feature in enumerate(FEATURES, start=1):
        row = 10 + index
        counts = feature_counts[feature["sheet_name"]]
        ws.cell(row, 2).value = index
        ws.cell(row, 3).value = feature["sheet_name"]
        ws.cell(row, 4).value = counts["Pass"]
        ws.cell(row, 5).value = counts["Fail"]
        ws.cell(row, 6).value = counts["Untested"]
        ws.cell(row, 7).value = counts["N/A"]
        ws.cell(row, 8).value = counts["Total"]

    total_pass = sum(value["Pass"] for value in feature_counts.values())
    total_fail = sum(value["Fail"] for value in feature_counts.values())
    total_untested = sum(value["Untested"] for value in feature_counts.values())
    total_na = sum(value["N/A"] for value in feature_counts.values())
    total_cases = total_pass + total_fail + total_untested + total_na
    executable_cases = total_cases - total_na
    executed_cases = total_pass + total_fail

    ws["D14"] = total_pass
    ws["E14"] = total_fail
    ws["F14"] = total_untested
    ws["G14"] = total_na
    ws["H14"] = total_cases
    ws["E16"] = round((executed_cases / executable_cases) * 100, 2) if executable_cases else 0
    ws["E17"] = round((total_pass / executable_cases) * 100, 2) if executable_cases else 0


def populate_detail_sheet(ws, feature: dict[str, Any], cases: list[dict[str, Any]]) -> None:
    keep_merges = {"B2:F2", "B3:F3", "B4:F4"}
    for merged_range in list(ws.merged_cells.ranges):
        if str(merged_range) not in keep_merges:
            ws.unmerge_cells(str(merged_range))

    clear_values(ws, 2, ws.max_row, 26)

    ws["A2"] = "Feature"
    ws["B2"] = feature["sheet_name"]
    ws["A3"] = "Test requirement"
    ws["B3"] = feature["requirement"]
    ws["A4"] = "Reference Document"
    ws["B4"] = "\n".join(feature["references"])
    ws["A5"] = "Pass"
    ws["B5"] = "Fail"
    ws["C5"] = "Untested"
    ws["D5"] = "N/A"
    ws["G5"] = "Number of Test cases"

    counts = count_results(cases)
    ws["A6"] = counts["Pass"]
    ws["B6"] = counts["Fail"]
    ws["C6"] = counts["Untested"]
    ws["D6"] = counts["N/A"]
    ws["G6"] = counts["Total"]
    ws["A9"] = "Integration/API flows (security, transformation, persistence, and response-contract checks)."

    grouped = defaultdict(list)
    for case in cases:
        grouped[case["section"]].append(case)

    row = 10
    for section in feature["sections"]:
        section_cases = grouped.get(section, [])
        if not section_cases:
            continue

        copy_row_style(ws, 10, row)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=26)
        ws.cell(row, 1).value = section
        row += 1

        for case in section_cases:
            copy_row_style(ws, 11, row)
            ws.cell(row, 1).value = f"[{case['id']}]"
            ws.cell(row, 2).value = case["description"]
            ws.cell(row, 3).value = case["procedure"]
            ws.cell(row, 4).value = case["expectedResults"]
            ws.cell(row, 5).value = detail_actual_text(case)
            ws.cell(row, 6).value = case.get("dependencies", "")
            ws.cell(row, 7).value = case.get("result", "Untested")
            row += 1

        row += 1


def generate_workbook() -> Path:
    cases = load_catalog()
    workbook = load_workbook(TEMPLATE_PATH)

    populate_cover(workbook["Cover"])
    populate_test_case_list(workbook["Test case list"], cases)
    populate_test_report(workbook["Test Report"], cases)

    feature_sheet_names = workbook.sheetnames[3:]
    base_feature_sheets = feature_sheet_names[:3]
    for old_name, feature in zip(base_feature_sheets, FEATURES):
        workbook[old_name].title = feature["sheet_name"]

    while len(workbook.sheetnames) > 6:
        del workbook[workbook.sheetnames[-1]]

    for feature in FEATURES:
        sheet = workbook[feature["sheet_name"]]
        feature_cases = [case for case in cases if case["feature"] == feature["sheet_name"]]
        populate_detail_sheet(sheet, feature, feature_cases)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    generate_workbook()
