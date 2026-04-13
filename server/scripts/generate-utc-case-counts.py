from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = REPO_ROOT / "docs" / "unit test final" / "Report5_Unit Test Case.xlsx"
OUTPUT_PATH = REPO_ROOT / "docs" / "unit test final" / "Report5_UTC_Case_Counts.json"


def main() -> None:
    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    report_sheet = workbook["Test Report"]

    document_code = report_sheet["B6"].value

    cases_by_ep = {}
    subtotal = None

    for row in range(12, report_sheet.max_row + 1):
        ep_code = report_sheet.cell(row, 2).value
        if ep_code == "Sub total":
            subtotal = {
                "passed": int(report_sheet.cell(row, 3).value or 0),
                "failed": int(report_sheet.cell(row, 4).value or 0),
                "untested": int(report_sheet.cell(row, 5).value or 0),
                "normal": int(report_sheet.cell(row, 6).value or 0),
                "abnormal": int(report_sheet.cell(row, 7).value or 0),
                "boundary": int(report_sheet.cell(row, 8).value or 0),
                "total": int(report_sheet.cell(row, 9).value or 0),
            }
            break

        if not ep_code:
            continue

        cases_by_ep[ep_code] = {
            "passed": int(report_sheet.cell(row, 3).value or 0),
            "failed": int(report_sheet.cell(row, 4).value or 0),
            "untested": int(report_sheet.cell(row, 5).value or 0),
            "normal": int(report_sheet.cell(row, 6).value or 0),
            "abnormal": int(report_sheet.cell(row, 7).value or 0),
            "boundary": int(report_sheet.cell(row, 8).value or 0),
            "total": int(report_sheet.cell(row, 9).value or 0),
        }

    payload = {
        "sourceWorkbook": str(WORKBOOK_PATH),
        "documentCode": document_code,
        "sheetCount": len(cases_by_ep),
        "subTotal": subtotal,
        "casesByEp": cases_by_ep,
    }

    OUTPUT_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
