from __future__ import annotations

import csv
import json
import re
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
DOCS_DIR = REPO_ROOT / "docs" / "unit test final"
WORKBOOK_PATH = DOCS_DIR / "Report5_Unit Test Case.xlsx"
AUDIT_CSV_PATH = DOCS_DIR / "Report5_Unit_Test_Jest_Audit.csv"
RUN_JSON_PATH = DOCS_DIR / "Report5_Unit_Test_Jest_Run_Result.json"

TRACE_CSV_PATH = DOCS_DIR / "Report5_UTC_Jest_Traceability.csv"
TRACE_MD_PATH = DOCS_DIR / "Report5_UTC_Jest_Traceability.md"
TRACE_JSON_PATH = DOCS_DIR / "Report5_UTC_Jest_Traceability.json"


def compact(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.lower())


def normalize_utc(utc_label: str) -> str:
    match = re.search(r"(\d+)$", utc_label)
    if not match:
        return utc_label.upper()
    return f"UTCID{int(match.group(1)):02d}"


def load_audit_rows() -> list[dict[str, str]]:
    with AUDIT_CSV_PATH.open(encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def load_assertions_by_file() -> dict[str, list[str]]:
    data = json.loads(RUN_JSON_PATH.read_text(encoding="utf-8"))
    assertions_by_file: dict[str, list[str]] = {}
    for test_result in data["testResults"]:
        file_path = str(Path(test_result["name"]))
        assertions_by_file[file_path] = [
            " > ".join(assertion.get("ancestorTitles", []) + [assertion.get("title", "")])
            for assertion in test_result.get("assertionResults", [])
        ]
    return assertions_by_file


def load_workbook_utc_columns():
    workbook = load_workbook(WORKBOOK_PATH, data_only=False)
    result: dict[str, list[str]] = {}
    for sheet in workbook.worksheets:
        if not sheet.title.startswith("EP-"):
            continue
        utc_columns: list[str] = []
        for column in range(1, sheet.max_column + 1):
            value = sheet.cell(9, column).value
            if isinstance(value, str) and value.upper().startswith("UTCID"):
                utc_columns.append(value.upper())
        result[sheet.title] = utc_columns
    return result


def choose_traceability(
    ep_code: str,
    function_name: str,
    titles: list[tuple[str, str]],
    utc_id: str,
) -> tuple[str, list[tuple[str, str]]]:
    utc_number = int(re.search(r"(\d+)$", utc_id).group(1))
    utc_pattern = re.compile(rf"UTC(?:ID)?{utc_number:02d}\b", re.IGNORECASE)
    ep_pattern = re.compile(rf"{re.escape(ep_code)}\s+UTC(?:ID)?{utc_number:02d}\b", re.IGNORECASE)
    function_key = compact(function_name)

    exact_ep = [(file_path, title) for file_path, title in titles if ep_pattern.search(title)]
    if exact_ep:
        return "exact_ep_title", exact_ep

    exact_fn = [
        (file_path, title)
        for file_path, title in titles
        if utc_pattern.search(title) and function_key and function_key in compact(title)
    ]
    if exact_fn:
        return "exact_function_title", exact_fn

    if titles:
        return "grouped_file_only", []

    return "no_supporting_title", []


def build_traceability_rows() -> list[dict[str, str]]:
    audit_rows = load_audit_rows()
    assertions_by_file = load_assertions_by_file()
    workbook_utc_columns = load_workbook_utc_columns()

    trace_rows: list[dict[str, str]] = []

    for audit_row in audit_rows:
        ep_code = audit_row["EP Code"]
        utc_columns = workbook_utc_columns.get(ep_code, [])
        matched_spec_files = [str(Path(item.strip())) for item in audit_row["Matched Spec Files"].split(" | ") if item.strip()]
        file_titles: list[tuple[str, str]] = []
        for spec_file in matched_spec_files:
            for title in assertions_by_file.get(spec_file, []):
                file_titles.append((spec_file, title))

        for utc_id in utc_columns:
            mapping_status, matches = choose_traceability(
                ep_code=ep_code,
                function_name=audit_row["Function Name"],
                titles=file_titles,
                utc_id=utc_id,
            )

            trace_rows.append(
                {
                    "EP Code": ep_code,
                    "Sheet": audit_row["Sheet"],
                    "Class Name": audit_row["Class Name"],
                    "Function Name": audit_row["Function Name"],
                    "UTC ID": normalize_utc(utc_id),
                    "Jest Status": audit_row["Jest Status"],
                    "Mapping Status": mapping_status,
                    "Matched Spec Files": " | ".join(matched_spec_files),
                    "Matched Assertion Titles": " || ".join(title for _, title in matches),
                    "Matched Assertion Files": " | ".join(sorted({file_path for file_path, _ in matches})),
                }
            )

    return trace_rows


def write_csv(rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "EP Code",
        "Sheet",
        "Class Name",
        "Function Name",
        "UTC ID",
        "Jest Status",
        "Mapping Status",
        "Matched Spec Files",
        "Matched Assertion Files",
        "Matched Assertion Titles",
    ]
    with TRACE_CSV_PATH.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_json(rows: list[dict[str, str]], summary: dict[str, int]) -> None:
    payload = {
        "sourceWorkbook": str(WORKBOOK_PATH),
        "sourceAuditCsv": str(AUDIT_CSV_PATH),
        "sourceRunJson": str(RUN_JSON_PATH),
        "summary": summary,
        "rows": rows,
    }
    TRACE_JSON_PATH.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")


def write_markdown(rows: list[dict[str, str]], summary: dict[str, int]) -> None:
    grouped_examples = [row for row in rows if row["Mapping Status"] == "grouped_file_only"][:10]
    exact_examples = [
        row
        for row in rows
        if row["Mapping Status"] in {"exact_ep_title", "exact_function_title"}
    ][:10]

    lines = [
        "# Report5 UTC -> Jest Traceability",
        "",
        f"- Total UTC columns in workbook: {summary['total_utc_columns']}",
        f"- Exact EP-title matches: {summary['exact_ep_title']}",
        f"- Exact function-title matches: {summary['exact_function_title']}",
        f"- Grouped file-only matches: {summary['grouped_file_only']}",
        f"- No supporting title rows: {summary['no_supporting_title']}",
        "",
        "## Meaning",
        "- `exact_ep_title`: a matched Jest assertion title includes the same `EP-xx` and UTC number.",
        "- `exact_function_title`: a matched Jest assertion title includes the same UTC number and the function name token, but not necessarily the same `EP-xx` code.",
        "- `grouped_file_only`: the EP/UTC column is backed by the mapped spec files, but no assertion title can be recovered exactly from the current artifacts.",
        "- `no_supporting_title`: no assertion titles were found in the mapped spec files.",
        "",
        "## Important Note",
        "This file is the strongest recoverable traceability from the current workbook and Jest artifacts. It is not safe to claim that all 1409 UTC columns have exact assertion-title matches, because many EP sheets are traced only at grouped file or batch level.",
        "",
        "## Exact Examples",
    ]

    for row in exact_examples:
        lines.append(
            f"- {row['EP Code']} {row['UTC ID']} {row['Function Name']} -> {row['Mapping Status']} -> `{row['Matched Assertion Titles'].split(' || ')[0]}`"
        )

    lines.extend(["", "## Grouped Examples"])

    for row in grouped_examples:
        lines.append(
            f"- {row['EP Code']} {row['UTC ID']} {row['Function Name']} -> grouped through `{row['Matched Spec Files'].split(' | ')[0]}`"
        )

    TRACE_MD_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    rows = build_traceability_rows()
    summary = Counter(row["Mapping Status"] for row in rows)
    summary["total_utc_columns"] = len(rows)
    for key in (
        "exact_ep_title",
        "exact_function_title",
        "grouped_file_only",
        "no_supporting_title",
    ):
        summary.setdefault(key, 0)

    write_csv(rows)
    write_json(rows, dict(summary))
    write_markdown(rows, dict(summary))

    print(TRACE_CSV_PATH)
    print(TRACE_MD_PATH)
    print(TRACE_JSON_PATH)
    print(dict(summary))


if __name__ == "__main__":
    main()
