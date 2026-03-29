#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.cell.cell import MergedCell

_SYNTHETIC_INPUT_SECTION_KEYS = {
    "action",
    "requeststatus",
    "currentstatus",
    "workflowstate",
    "statetransition",
    "requeststate",
}


@dataclass(frozen=True)
class FunctionSheetLayout:
    template_sheet_name: str
    case_start_col: int
    utc_row: int
    condition_row: int
    confirm_row: int
    result_row: int
    passed_cell: str
    failed_cell: str
    untested_cell: str
    n_count_cell: str
    a_count_cell: str
    b_count_cell: str
    total_cell: str


def _read_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _resolve_default_template_path() -> Path:
    candidate_dirs = [
        Path.cwd() / "docs" / "Unit Testing Excel",
        Path.cwd() / "docs" / "template" / "unit-test",
    ]
    versioned: List[Tuple[Tuple[int, ...], Path]] = []

    for directory in candidate_dirs:
        if not directory.exists():
            continue
        for workbook in directory.glob("Report5_Unit Test Case_v*.xlsx"):
            match = re.search(r"_v(\d+(?:\.\d+)*)\.xlsx$", workbook.name)
            if not match:
                continue
            version = tuple(int(part) for part in match.group(1).split("."))
            versioned.append((version, workbook))

    if versioned:
        versioned.sort(key=lambda item: item[0])
        return versioned[-1][1]

    raise SystemExit("Could not find a Report5_Unit Test Case template workbook.")


def _find_cell(ws: openpyxl.worksheet.worksheet.Worksheet, needle: str) -> Optional[Tuple[int, int]]:
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value == needle:
                return (cell.row, cell.column)
    return None


def _find_row_col_a(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    needle: str,
) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == needle:
            return row
    return None


def _cell_below_label(ws: openpyxl.worksheet.worksheet.Worksheet, label: str) -> str:
    pos = _find_cell(ws, label)
    if not pos:
        raise ValueError(f"Template sheet missing required label: {label}")
    row, col = pos
    return openpyxl.utils.get_column_letter(col) + str(row + 1)


def _utc_info(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> Optional[Tuple[int, int, int]]:
    pos = _find_cell(ws, "UTCID01")
    if not pos:
        return None
    utc_row, utc_col = pos
    last_col = utc_col
    for col in range(utc_col, ws.max_column + 1):
        value = ws.cell(utc_row, col).value
        if isinstance(value, str) and value.startswith("UTCID"):
            last_col = col
    return utc_row, utc_col, last_col


def _pick_template_sheet_name(wb: openpyxl.Workbook, preferred: Optional[str]) -> str:
    if preferred:
        if preferred not in wb.sheetnames:
            raise SystemExit(f"Template sheet not found: {preferred}")
        return preferred

    helper_sheets = {"Guideline", "Cover", "Function List", "Test Report", "Statistics"}
    best_name: Optional[str] = None
    best_utc_count = -1
    best_rows = -1

    for name in wb.sheetnames:
        if name in helper_sheets:
            continue
        ws = wb[name]
        utc = _utc_info(ws)
        if not utc:
            continue
        if not _find_row_col_a(ws, "Condition") or not _find_row_col_a(ws, "Confirm") or not _find_row_col_a(ws, "Result"):
            continue
        utc_row, utc_col, last_col = utc
        utc_count = 0
        for col in range(utc_col, last_col + 1):
            value = ws.cell(utc_row, col).value
            if isinstance(value, str) and value.startswith("UTCID"):
                utc_count += 1
        if (utc_count, ws.max_row) > (best_utc_count, best_rows):
            best_name = name
            best_utc_count = utc_count
            best_rows = ws.max_row

    if not best_name:
        raise SystemExit("No suitable function sheet template found.")
    return best_name


def _detect_layout(wb: openpyxl.Workbook, template_sheet_name: str) -> FunctionSheetLayout:
    ws = wb[template_sheet_name]
    utc = _utc_info(ws)
    if not utc:
        raise SystemExit(f"Template function sheet '{template_sheet_name}' has no UTCID01 row.")
    utc_row, case_start_col, _ = utc

    condition_row = _find_row_col_a(ws, "Condition")
    confirm_row = _find_row_col_a(ws, "Confirm")
    result_row = _find_row_col_a(ws, "Result")
    if not condition_row or not confirm_row or not result_row:
        raise SystemExit("Template function sheet must include Condition/Confirm/Result rows in column A.")

    n_a_b_pos = _find_cell(ws, "N/A/B")
    total_pos = _find_cell(ws, "Total Test Cases")
    if not n_a_b_pos or not total_pos:
        raise SystemExit("Template function sheet missing N/A/B or Total Test Cases labels.")

    return FunctionSheetLayout(
        template_sheet_name=template_sheet_name,
        case_start_col=case_start_col,
        utc_row=utc_row,
        condition_row=condition_row,
        confirm_row=confirm_row,
        result_row=result_row,
        passed_cell=_cell_below_label(ws, "Passed"),
        failed_cell=_cell_below_label(ws, "Failed"),
        untested_cell=_cell_below_label(ws, "Untested"),
        n_count_cell=openpyxl.utils.get_column_letter(n_a_b_pos[1]) + str(n_a_b_pos[0] + 1),
        a_count_cell=openpyxl.utils.get_column_letter(n_a_b_pos[1] + 1) + str(n_a_b_pos[0] + 1),
        b_count_cell=openpyxl.utils.get_column_letter(n_a_b_pos[1] + 2) + str(n_a_b_pos[0] + 1),
        total_cell=openpyxl.utils.get_column_letter(total_pos[1]) + str(total_pos[0] + 1),
    )


def _clear_range(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def _sanitize_sheet_name(name: str) -> str:
    name = (name or "").strip()
    if not name:
        return "Function"
    name = re.sub(r"[:\\/\?\*\[\]]", "_", name)
    return name[:31]


def _make_unique_sheet_name(existing: Iterable[str], desired: str) -> str:
    existing_set = set(existing)
    if desired not in existing_set:
        return desired
    base = desired[:28] if len(desired) > 28 else desired
    idx = 2
    while True:
        candidate = f"{base}-{idx}"
        if candidate not in existing_set:
            return candidate
        idx += 1


def _normalize_scalar(value: Any) -> Any:
    if value is None:
        return "null"
    return value


def _as_list(value: Any) -> List[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def _str_key(value: Any) -> str:
    if isinstance(value, str):
        return value
    try:
        return json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    except TypeError:
        return str(value)


def _normalize_input_section_key(value: Any) -> str:
    return re.sub(r"[\s_-]+", "", str(value or "").lower())


def _assert_real_input_name(function_name: str, input_name: str) -> None:
    if _normalize_input_section_key(input_name) not in _SYNTHETIC_INPUT_SECTION_KEYS:
        return
    raise SystemExit(
        f"Function '{function_name}' case input '{input_name}' is a synthetic worksheet section. "
        "Use real request parameters instead."
    )


def _write_marks(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col_start: int,
    active_case_indexes: Sequence[int],
) -> None:
    for idx in active_case_indexes:
        ws.cell(row, col_start + idx).value = "O"


def _set_header_value(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    label: str,
    offset_cols: int,
    value: Any,
    search_max_row: int = 10,
    search_max_col: int = 20,
) -> None:
    for row in range(1, min(search_max_row, ws.max_row) + 1):
        for col in range(1, min(search_max_col, ws.max_column) + 1):
            if ws.cell(row, col).value == label:
                ws.cell(row, col + offset_cols).value = value
                return


def _flatten_assertions(results_payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    flattened: List[Dict[str, Any]] = []
    for suite in results_payload.get("testResults", []):
        for assertion in suite.get("assertionResults", []):
            flattened.append(
                {
                    "status": assertion.get("status"),
                    "title": assertion.get("title", ""),
                    "fullName": assertion.get("fullName", ""),
                }
            )
    return flattened


def _match_assertions(assertions: List[Dict[str, Any]], test_names: Sequence[str]) -> List[Dict[str, Any]]:
    matched: List[Dict[str, Any]] = []
    for test_name in test_names:
        for assertion in assertions:
            full_name = assertion["fullName"]
            title = assertion["title"]
            if full_name == test_name or title == test_name or full_name.endswith(test_name):
                matched.append(assertion)
    return matched


def _apply_evidence(functions: List[Dict[str, Any]], assertions: List[Dict[str, Any]], default_date: str) -> None:
    for function in functions:
        for case in function.get("cases", []):
            test_names = case.get("test_names") or []
            if case.get("test_name"):
                test_names = list(test_names) + [case["test_name"]]
            matches = _match_assertions(assertions, test_names)
            if not matches:
                case["status"] = "U"
                case["executed_date"] = ""
                if not case.get("logs") and test_names:
                    case["logs"] = [f'Jest: {test_names[0]}']
                continue

            statuses = {match["status"] for match in matches}
            case["status"] = "F" if "failed" in statuses else "P"
            case["executed_date"] = default_date
            if not case.get("logs") and test_names:
                case["logs"] = [f'Jest: {test_names[0]}']


def _populate_function_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    layout: FunctionSheetLayout,
    function: Dict[str, Any],
) -> None:
    cases = function.get("cases") or []
    _set_header_value(ws, "Function Code", 2, function.get("function_code") or "")
    _set_header_value(ws, "Function Name", 6, function.get("function_name") or "")
    _set_header_value(ws, "Created By", 2, function.get("created_by") or "")
    _set_header_value(ws, "Executed By", 6, function.get("executed_by") or function.get("created_by") or "")
    _set_header_value(ws, "Lines  of code", 2, function.get("loc") or "")
    _set_header_value(ws, "Test requirement", 2, function.get("test_requirement") or "")

    end_row = layout.result_row + 3
    _clear_range(ws, layout.utc_row, end_row, 1, ws.max_column)

    ws.cell(layout.condition_row, 1).value = "Condition"
    ws.cell(layout.condition_row, 2).value = "Precondition "
    ws.cell(layout.confirm_row, 1).value = "Confirm"
    ws.cell(layout.confirm_row, 2).value = "Return"
    ws.cell(layout.result_row, 1).value = "Result"
    ws.cell(layout.result_row, 2).value = "Type(N : Normal, A : Abnormal, B : Boundary)"
    ws.cell(layout.result_row + 1, 2).value = "Passed/Failed"
    ws.cell(layout.result_row + 2, 2).value = "Executed Date"
    ws.cell(layout.result_row + 3, 2).value = "Defect ID"

    for idx in range(len(cases)):
        ws.cell(layout.utc_row, layout.case_start_col + idx).value = f"UTCID{idx + 1:02d}"

    preconditions_map: Dict[str, List[int]] = {}
    inputs_map: Dict[str, Dict[str, Dict[str, Any]]] = {}
    returns_map: Dict[str, List[int]] = {}
    exceptions_map: Dict[str, List[int]] = {}
    logs_map: Dict[str, List[int]] = {}

    for idx, case in enumerate(cases):
        for precondition in _as_list(case.get("preconditions")):
            text = str(precondition)
            preconditions_map.setdefault(text, []).append(idx)

        inputs = case.get("inputs") or {}
        if not isinstance(inputs, dict):
            raise SystemExit(f"Function '{function.get('function_name')}' case inputs must be an object.")

        for input_name, raw_value in inputs.items():
            _assert_real_input_name(str(function.get("function_name") or ""), str(input_name))
            inputs_map.setdefault(input_name, {})
            for value in _as_list(raw_value):
                normalized = _normalize_scalar(value)
                key = _str_key(normalized)
                inputs_map[input_name].setdefault(key, {"display": normalized, "cases": []})
                inputs_map[input_name][key]["cases"].append(idx)

        for returned in _as_list(case.get("returns")):
            returns_map.setdefault(str(returned), []).append(idx)

        for exception in _as_list(case.get("exceptions")):
            exceptions_map.setdefault(str(exception), []).append(idx)

        for log_message in _as_list(case.get("logs")):
            logs_map.setdefault(str(log_message), []).append(idx)

    condition_capacity = layout.confirm_row - layout.condition_row - 1
    condition_required = len(preconditions_map) + sum(1 + len(value) for value in inputs_map.values())
    if condition_required > condition_capacity:
        raise SystemExit(
            f"Function '{function.get('function_name')}' needs {condition_required} condition rows but template only has {condition_capacity}."
        )

    confirm_capacity = layout.result_row - layout.confirm_row - 1
    confirm_required = (
        len(returns_map)
        + (0 if not exceptions_map else 1 + len(exceptions_map))
        + (0 if not logs_map else 1 + len(logs_map))
    )
    if confirm_required > confirm_capacity:
        raise SystemExit(
            f"Function '{function.get('function_name')}' needs {confirm_required} confirm rows but template only has {confirm_capacity}."
        )

    row = layout.condition_row + 1
    for text, case_indexes in preconditions_map.items():
        ws.cell(row, 4).value = text
        _write_marks(ws, row, layout.case_start_col, case_indexes)
        row += 1

    for input_name, values_map in inputs_map.items():
        ws.cell(row, 2).value = input_name
        row += 1
        for entry in values_map.values():
            ws.cell(row, 4).value = entry["display"]
            _write_marks(ws, row, layout.case_start_col, entry["cases"])
            row += 1

    row = layout.confirm_row + 1
    for text, case_indexes in returns_map.items():
        ws.cell(row, 4).value = text
        _write_marks(ws, row, layout.case_start_col, case_indexes)
        row += 1

    if exceptions_map:
        ws.cell(row, 2).value = "Exception"
        row += 1
        for text, case_indexes in exceptions_map.items():
            ws.cell(row, 4).value = text
            _write_marks(ws, row, layout.case_start_col, case_indexes)
            row += 1

    if logs_map:
        ws.cell(row, 2).value = "Log message"
        row += 1
        for text, case_indexes in logs_map.items():
            ws.cell(row, 4).value = text
            _write_marks(ws, row, layout.case_start_col, case_indexes)
            row += 1

    for idx, case in enumerate(cases):
        ws.cell(layout.result_row, layout.case_start_col + idx).value = (case.get("type") or "").strip().upper()
        ws.cell(layout.result_row + 1, layout.case_start_col + idx).value = (case.get("status") or "").strip().upper()
        ws.cell(layout.result_row + 2, layout.case_start_col + idx).value = case.get("executed_date") or ""
        ws.cell(layout.result_row + 3, layout.case_start_col + idx).value = case.get("defect_id") or ""

    passed_count = sum(1 for case in cases if (case.get("status") or "").strip().upper() == "P")
    failed_count = sum(1 for case in cases if (case.get("status") or "").strip().upper() == "F")
    untested_count = max(len(cases) - passed_count - failed_count, 0)
    n_count = sum(1 for case in cases if (case.get("type") or "").strip().upper() == "N")
    a_count = sum(1 for case in cases if (case.get("type") or "").strip().upper() == "A")
    b_count = sum(1 for case in cases if (case.get("type") or "").strip().upper() == "B")

    ws[layout.passed_cell] = passed_count
    ws[layout.failed_cell] = failed_count
    ws[layout.untested_cell] = untested_count
    ws[layout.n_count_cell] = n_count
    ws[layout.a_count_cell] = a_count
    ws[layout.b_count_cell] = b_count
    ws[layout.total_cell] = len(cases)


def _update_cover(wb: openpyxl.Workbook, meta: Dict[str, Any]) -> None:
    if "Cover" not in wb.sheetnames:
        return
    ws = wb["Cover"]
    if meta.get("project_name"):
        ws["B4"].value = meta["project_name"]
    if meta.get("project_code"):
        ws["B5"].value = meta["project_code"]
    if meta.get("creator"):
        ws["F4"].value = meta["creator"]
    if meta.get("reviewer"):
        ws["F5"].value = meta["reviewer"]
    if meta.get("issue_date"):
        ws["F6"].value = meta["issue_date"]
    if meta.get("version"):
        ws["F7"].value = meta["version"]


def _update_function_list(wb: openpyxl.Workbook, function: Dict[str, Any], sheet_name: str) -> None:
    if "Function List" not in wb.sheetnames:
        return
    ws = wb["Function List"]
    _clear_range(ws, 11, max(ws.max_row, 30), 1, 7)
    ws.cell(11, 1).value = 1
    ws.cell(11, 2).value = function.get("class_name") or ""
    ws.cell(11, 3).value = function.get("function_name") or ""
    ws.cell(11, 4).value = function.get("function_code") or ""
    ws.cell(11, 5).value = sheet_name
    ws.cell(11, 6).value = function.get("description") or ""
    ws.cell(11, 7).value = function.get("precondition") or ""


def _escape_sheet_ref(sheet_name: str) -> str:
    return sheet_name.replace("'", "''")


def _update_test_report(
    wb: openpyxl.Workbook,
    function: Dict[str, Any],
    sheet_name: str,
    layout: FunctionSheetLayout,
) -> None:
    if "Test Report" not in wb.sheetnames:
        return
    ws = wb["Test Report"]
    _clear_range(ws, 12, max(ws.max_row, 30), 1, 9)
    sheet_ref = _escape_sheet_ref(sheet_name)
    ws.cell(12, 1).value = 1
    ws.cell(12, 2).value = function.get("function_code") or ""
    ws.cell(12, 3).value = f"='{sheet_ref}'!{layout.passed_cell}"
    ws.cell(12, 4).value = f"='{sheet_ref}'!{layout.failed_cell}"
    ws.cell(12, 5).value = f"='{sheet_ref}'!{layout.untested_cell}"
    ws.cell(12, 6).value = f"='{sheet_ref}'!{layout.n_count_cell}"
    ws.cell(12, 7).value = f"='{sheet_ref}'!{layout.a_count_cell}"
    ws.cell(12, 8).value = f"='{sheet_ref}'!{layout.b_count_cell}"
    ws.cell(12, 9).value = f"='{sheet_ref}'!{layout.total_cell}"


def _clone_sheet_format_only(
    workbook: openpyxl.Workbook,
    source_name: str,
    desired_name: str,
) -> openpyxl.worksheet.worksheet.Worksheet:
    source = workbook[source_name]
    cloned = workbook.copy_worksheet(source)
    cloned.title = desired_name
    return cloned


def _prune_function_sheets(
    workbook: openpyxl.Workbook,
    template_sheet_name: str,
    active_sheet_name: str,
) -> None:
    helper_sheets = {"Guideline", "Cover", "Function List", "Test Report", "Statistics"}
    for sheet_name in list(workbook.sheetnames):
        if sheet_name in helper_sheets:
            continue
        if sheet_name == active_sheet_name:
            continue
        workbook.remove(workbook[sheet_name])


def _build_output_file_name(function: Dict[str, Any], version: str, created_by: str) -> str:
    if function.get("output_file"):
        return str(function["output_file"])
    function_code = function.get("function_code") or "FN"
    return f"Report5_Unit Test Case_{version}_{function_code}_{created_by}.xlsx"


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Generate per-function controller UTC workbooks from a manifest + Jest JSON results.")
    parser.add_argument("--manifest", required=True, help="Path to the manifest JSON.")
    parser.add_argument("--results", required=True, help="Path to the Jest JSON result file.")
    parser.add_argument("--output-dir", required=True, help="Directory for generated workbooks.")
    parser.add_argument("--template", default=str(_resolve_default_template_path()), help="Path to the template workbook.")
    parser.add_argument("--template-function-sheet", help="Optional function sheet to clone from the template.")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing output files.")
    args = parser.parse_args(argv)

    manifest_path = Path(args.manifest)
    results_path = Path(args.results)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    manifest = _read_json(manifest_path)
    results_payload = _read_json(results_path)
    functions = manifest.get("functions") or []
    if not isinstance(functions, list) or not functions:
        raise SystemExit("Manifest must contain a non-empty functions[] array.")

    meta = manifest.get("meta") or {}
    template_path = Path(args.template)
    if not template_path.exists():
        raise SystemExit(f"Template workbook not found: {template_path}")

    assertions = _flatten_assertions(results_payload)
    default_date = str(meta.get("executed_date") or meta.get("issue_date") or "")
    _apply_evidence(functions, assertions, default_date)

    template_workbook = openpyxl.load_workbook(template_path)
    template_sheet_name = _pick_template_sheet_name(template_workbook, args.template_function_sheet)
    layout = _detect_layout(template_workbook, template_sheet_name)

    generated_paths: List[Path] = []
    for function in functions:
        workbook = openpyxl.load_workbook(template_path)
        template_sheet_name = _pick_template_sheet_name(workbook, args.template_function_sheet)
        layout = _detect_layout(workbook, template_sheet_name)

        desired_name = _sanitize_sheet_name(
            str(function.get("sheet_name") or function.get("function_name") or function.get("function_code") or "Function")
        )
        actual_name = template_sheet_name
        if desired_name != template_sheet_name:
            desired_name = _make_unique_sheet_name(
                [name for name in workbook.sheetnames if name != template_sheet_name],
                desired_name,
            )
            workbook[template_sheet_name].title = desired_name
            actual_name = desired_name

        _prune_function_sheets(workbook, template_sheet_name, actual_name)
        active_sheet = workbook[actual_name]
        active_sheet.sheet_state = "visible"

        _populate_function_sheet(active_sheet, layout, function)
        _update_cover(workbook, meta)
        _update_function_list(workbook, function, actual_name)
        _update_test_report(workbook, function, actual_name, layout)
        workbook.defined_names.clear()

        output_name = _build_output_file_name(
            function,
            str(meta.get("version") or "v2.4"),
            str(function.get("created_by") or meta.get("creator") or "author"),
        )
        output_path = output_dir / output_name
        if output_path.exists() and not args.overwrite:
            raise SystemExit(f"Output already exists: {output_path}")

        workbook.save(output_path)
        generated_paths.append(output_path.resolve())

    for path in generated_paths:
        print(str(path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
