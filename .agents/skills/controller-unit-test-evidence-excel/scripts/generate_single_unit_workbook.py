#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import math
import re
import sys
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidationList
from openpyxl.worksheet.worksheet import Worksheet

HELPER_SHEETS = {"Guideline", "Cover", "Function List", "Test Report"}


@dataclass(frozen=True)
class FunctionSheetLayout:
    utc_row: int
    case_start_col: int
    last_case_col: int
    condition_header_row: int
    precondition_label_col: int
    content_col: int
    passed_cell: str
    failed_cell: str
    untested_cell: str
    n_count_cell: str
    a_count_cell: str
    b_count_cell: str
    total_cell: str


def compact_test_requirement(text: str) -> str:
    text = " ".join(text.split())
    match = re.match(r"Verify\s+(?:[A-Za-z0-9_]+Controller\.)?([A-Za-z0-9_]+)\s+handles.+for\s+(EP-\d+)\.$", text)
    if match:
        method_name, code = match.groups()
        return f"Verify {method_name}: valid, edge, error, and auth cases for {code}."
    return text


def compact_case_text(text: Any, section: str) -> Any:
    if not isinstance(text, str):
        return text

    compacted = " ".join(text.replace("\n", " ").split()).strip().strip('"')
    compacted = re.sub(r"\b([A-Za-z0-9_]+Controller)\.([A-Za-z0-9_]+)\b", r"\2", compacted)

    replacements = [
        ("Authenticated ", ""),
        (" user invokes ", " user calls "),
        ("Dependencies required by ", ""),
        (" dependencies are available", " available"),
        ("Caller omits the optional ", "Omit "),
        (" route is protected by authentication middleware", " auth middleware enabled"),
        ("Caller is authenticated as a ", "Caller role is "),
        (" role", ""),
        ("returns ", "Returns "),
        (" payload", ""),
        (' for range = "', ' for "'),
        (' for default range = "', ' default "'),
        (' for normalized range = "', ' normalized "'),
        ("Handled edge case:", "Edge:"),
        ("Validation outcome:", "Validation:"),
        ("Security outcome:", "Security:"),
        ("Success:", "Success:"),
        (" when it is omitted", ""),
        (" when it is empty", " when empty"),
        ("declares JwtAuthGuard on the ", "JwtAuthGuard on "),
        (" method.", "."),
        (" method", ""),
        ("restricts the ", ""),
        (" to ADMIN role metadata.", " to ADMIN only."),
        ("only ADMIN can access ", "ADMIN only "),
        ("propagates the business exception for an unsupported range.", "rejects unsupported range."),
        ("propagates a temporary analytics outage.", "reports analytics outage."),
        ("bubbles up unexpected repository failures.", "bubbles repository failure."),
        ("returns overview payload for ", "returns overview for "),
        ("returns audit log export file with Content-Type = ", "returns export file "),
        ("with XLSX attachment headers", "with XLSX headers"),
        ("with the expected attachment headers", "with expected headers"),
        ("Success: returns overview for a valid range input.", "Success: valid overview."),
    ]
    for source, target in replacements:
        compacted = compacted.replace(source, target)

    compacted = re.sub(r"\bOverview route\b", "Overview", compacted)
    compacted = re.sub(r"\bunsupported range aliases\b", "unsupported range", compacted)
    compacted = re.sub(r"\s+to the default range\b", " to default", compacted)
    compacted = re.sub(r"\b403 Forbidden: ADMIN only \w+\b", "403 Forbidden: ADMIN only", compacted)
    compacted = re.sub(r"\b401 Unauthorized\b", "401 Unauthorized", compacted)
    compacted = re.sub(r"\bReturns overview for \"([^\"]+)\"?$", r"Overview \1", compacted)
    compacted = re.sub(r"\bReturns overview default \"([^\"]+)\"?$", r"Overview default \1", compacted)
    compacted = re.sub(r"\bReturns overview normalized \"([^\"]+)\"?$", r"Overview normalized \1", compacted)
    compacted = re.sub(r"\breturns overview for valid range input\.\b", "returns overview for valid range.", compacted, flags=re.IGNORECASE)
    compacted = re.sub(r"\bdefaults the range input to 30d\b", "defaults range to 30d", compacted, flags=re.IGNORECASE)
    compacted = re.sub(r"\bnormalizes an unsupported range input to 30d\b", "normalizes unsupported range to 30d", compacted, flags=re.IGNORECASE)
    compacted = re.sub(r"^Success: returns (.+?) for a valid range input\.$", r"Success: valid \1.", compacted)
    compacted = re.sub(r"^Success: returns .+valid range input\.$", "Success: valid response.", compacted)
    compacted = re.sub(r"^Edge: defaults range to 30d\.$", "Edge: omitted range -> 30d.", compacted)
    compacted = re.sub(r"^Edge: normalizes unsupported range to 30d\.$", "Edge: unsupported range -> 30d.", compacted)
    compacted = re.sub(r"^Validation: rejects unsupported range\.$", "Validation: unsupported range.", compacted)
    compacted = re.sub(r"^Validation: reports analytics outage\.$", "Validation: analytics outage.", compacted)
    compacted = re.sub(r"^Validation: bubbles repository failure\.$", "Validation: repository failure.", compacted)
    compacted = re.sub(r"^Security: JwtAuthGuard on .+\.$", "Security: JwtAuthGuard.", compacted)
    compacted = re.sub(r"^Security: .+ ADMIN only\.$", "Security: ADMIN only.", compacted)
    compacted = re.sub(r"^Security: .+ ADMIN metadata\.$", "Security: ADMIN only.", compacted)

    if compacted.count('"') % 2 == 1:
        compacted += '"'

    if section == "precondition":
        compacted = compacted.replace(" user calls ", " calls ")
    if section == "exception":
        compacted = compacted.replace("403 Forbidden: ADMIN only ", "403 Forbidden: ADMIN only ")
    if section == "log":
        compacted = compacted.replace("Success: Returns ", "Success: returns ")
        compacted = compacted.replace("Edge:", "Edge:")
        compacted = compacted.replace("Validation:", "Validation:")
        compacted = compacted.replace("Security:", "Security:")

    if len(compacted) > 72:
        sentence_break = re.split(r"(?<=[\.:;])\s+", compacted)
        if sentence_break:
            compacted = sentence_break[0]
    if len(compacted) > 72:
        compacted = compacted[:69].rstrip() + "..."

    return compacted


def read_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def find_cell(ws: Worksheet, needle: str) -> Optional[Tuple[int, int]]:
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value == needle:
                return cell.row, cell.column
    return None


def find_row_in_col_a(ws: Worksheet, needle: str) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == needle:
            return row
    return None


def pick_template_sheet_name(wb: openpyxl.Workbook) -> str:
    best_name: Optional[str] = None
    best_count = -1

    for name in wb.sheetnames:
        if name in HELPER_SHEETS:
            continue
        ws = wb[name]
        if not find_row_in_col_a(ws, "Condition") or not find_cell(ws, "UTCID01"):
            continue

        count = 0
        for cell in ws[find_cell(ws, "UTCID01")[0]]:
            if isinstance(cell.value, str) and cell.value.startswith("UTCID"):
                count += 1

        if count > best_count:
            best_name = name
            best_count = count

    if not best_name:
        raise SystemExit("Could not find a function-sheet template in the workbook.")
    return best_name


def detect_layout(ws: Worksheet) -> FunctionSheetLayout:
    utc_pos = find_cell(ws, "UTCID01")
    passed_pos = find_cell(ws, "Passed")
    failed_pos = find_cell(ws, "Failed")
    untested_pos = find_cell(ws, "Untested")
    n_a_b_pos = find_cell(ws, "N/A/B")
    total_pos = find_cell(ws, "Total Test Cases")
    condition_row = find_row_in_col_a(ws, "Condition")
    precondition_pos = find_cell(ws, "Precondition ")

    if not all([utc_pos, passed_pos, failed_pos, untested_pos, n_a_b_pos, total_pos, condition_row, precondition_pos]):
        raise SystemExit("Unit-test template is missing the required labels.")

    utc_row, case_start_col = utc_pos
    last_case_col = case_start_col
    for col in range(case_start_col, ws.max_column + 1):
        value = ws.cell(utc_row, col).value
        if isinstance(value, str) and value.startswith("UTCID"):
            last_case_col = col

    return FunctionSheetLayout(
        utc_row=utc_row,
        case_start_col=case_start_col,
        last_case_col=last_case_col,
        condition_header_row=condition_row,
        precondition_label_col=precondition_pos[1],
        content_col=case_start_col - 2,
        passed_cell=f"{openpyxl.utils.get_column_letter(passed_pos[1])}{passed_pos[0] + 1}",
        failed_cell=f"{openpyxl.utils.get_column_letter(failed_pos[1])}{failed_pos[0] + 1}",
        untested_cell=f"{openpyxl.utils.get_column_letter(untested_pos[1])}{untested_pos[0] + 1}",
        n_count_cell=f"{openpyxl.utils.get_column_letter(n_a_b_pos[1])}{n_a_b_pos[0] + 1}",
        a_count_cell=f"{openpyxl.utils.get_column_letter(n_a_b_pos[1] + 1)}{n_a_b_pos[0] + 1}",
        b_count_cell=f"{openpyxl.utils.get_column_letter(n_a_b_pos[1] + 2)}{n_a_b_pos[0] + 1}",
        total_cell=f"{openpyxl.utils.get_column_letter(total_pos[1])}{total_pos[0] + 1}",
    )


def copy_row_style(source_ws: Worksheet, source_row: int, target_ws: Worksheet, target_row: int) -> None:
    source_dimension = source_ws.row_dimensions[source_row]
    target_dimension = target_ws.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed

    for col in range(1, source_ws.max_column + 1):
        source = source_ws.cell(source_row, col)
        target = target_ws.cell(target_row, col)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)
        target.fill = copy(source.fill)
        target.font = copy(source.font)
        target.border = copy(source.border)


def copy_row_merges(source_ws: Worksheet, source_row: int, target_ws: Worksheet, target_row: int) -> None:
    for merged_range in list(target_ws.merged_cells.ranges):
        if merged_range.min_row == target_row == merged_range.max_row:
            target_ws.unmerge_cells(str(merged_range))

    for merged_range in source_ws.merged_cells.ranges:
        if merged_range.min_row == source_row == merged_range.max_row:
            start_col = openpyxl.utils.get_column_letter(merged_range.min_col)
            end_col = openpyxl.utils.get_column_letter(merged_range.max_col)
            target_ws.merge_cells(f"{start_col}{target_row}:{end_col}{target_row}")


def copy_sheet_view(source_ws: Worksheet, target_ws: Worksheet) -> None:
    target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale
    target_ws.sheet_view.zoomScaleNormal = source_ws.sheet_view.zoomScaleNormal
    target_ws.sheet_view.zoomScalePageLayoutView = source_ws.sheet_view.zoomScalePageLayoutView
    target_ws.sheet_view.zoomScaleSheetLayoutView = source_ws.sheet_view.zoomScaleSheetLayoutView
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
    target_ws.sheet_view.showRowColHeaders = source_ws.sheet_view.showRowColHeaders
    target_ws.sheet_view.view = source_ws.sheet_view.view


def clear_sheet_artifacts(ws: Worksheet) -> None:
    ws.data_validations = DataValidationList()
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.comment = None


def trim_function_sheet(ws: Worksheet, last_row: int) -> None:
    if ws.max_row > last_row:
        trailing_count = ws.max_row - last_row
        ws.delete_rows(last_row + 1, trailing_count)
        for row in range(last_row + 1, last_row + trailing_count + 1):
            ws.row_dimensions[row].hidden = True

    for col in range(openpyxl.utils.column_index_from_string("U"), ws.max_column + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].hidden = True


def estimate_wrapped_height(value: Any, width_hint: float, base_height: Optional[float]) -> float:
    text = "" if value is None else str(value)
    explicit_lines = max(1, text.count("\n") + 1)
    chars_per_line = max(10, int(width_hint * 1.0))
    wrapped_lines = max(1, math.ceil(len(text) / chars_per_line))
    line_count = max(explicit_lines, wrapped_lines)
    baseline = base_height or 15
    return baseline * line_count


def merged_width_hint(ws: Worksheet, row: int, col: int) -> float:
    for merged_range in ws.merged_cells.ranges:
        if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
            total = 0.0
            for index in range(merged_range.min_col, merged_range.max_col + 1):
                total += ws.column_dimensions[openpyxl.utils.get_column_letter(index)].width or 8.43
            return total
    return ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width or 26


def apply_content_font(cell: openpyxl.cell.cell.Cell) -> None:
    font = copy(cell.font)
    font.name = "Calibri"
    font.size = 12
    cell.font = font


def write_value(
    ws: Worksheet,
    cell_ref: str,
    value: Any,
    *,
    content_font: bool = True,
    wrap_text: Optional[bool] = None,
) -> None:
    cell = ws[cell_ref]
    cell.value = value
    if content_font:
        apply_content_font(cell)
    if wrap_text is not None:
        alignment = copy(cell.alignment)
        alignment.wrap_text = wrap_text
        alignment.shrinkToFit = False
        cell.alignment = alignment
        if wrap_text:
            width_hint = merged_width_hint(ws, cell.row, cell.column)
            ws.row_dimensions[cell.row].height = estimate_wrapped_height(value, width_hint, ws.row_dimensions[cell.row].height)


def write_row_label(ws: Worksheet, row: int, layout: FunctionSheetLayout, value: str) -> None:
    cell = ws.cell(row, layout.precondition_label_col)
    cell.value = value
    apply_content_font(cell)


def write_content(ws: Worksheet, row: int, layout: FunctionSheetLayout, value: Any, *, wrap_text: bool = False) -> None:
    cell = ws.cell(row, layout.content_col)
    cell.value = value
    apply_content_font(cell)
    alignment = copy(cell.alignment)
    alignment.horizontal = "left"
    alignment.vertical = "top"
    alignment.wrap_text = wrap_text
    alignment.shrinkToFit = False
    cell.alignment = alignment
    if wrap_text:
        width_hint = merged_width_hint(ws, row, layout.content_col)
        ws.row_dimensions[row].height = estimate_wrapped_height(value, width_hint, ws.row_dimensions[row].height)


def write_case_mark(ws: Worksheet, row: int, layout: FunctionSheetLayout, case_indexes: Sequence[int]) -> None:
    for index in case_indexes:
        cell = ws.cell(row, layout.case_start_col + index)
        cell.value = "O"
        apply_content_font(cell)


def normalize_scalar(value: Any) -> Any:
    if value is None:
        return "null"
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False)
    return value


def as_list(value: Any) -> List[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def grouped_case_rows(function: Dict[str, Any]):
    preconditions: Dict[str, List[int]] = {}
    inputs: Dict[str, Dict[str, Dict[str, Any]]] = {}
    returns: Dict[str, List[int]] = {}
    exceptions: Dict[str, List[int]] = {}
    logs: Dict[str, List[int]] = {}

    for idx, case in enumerate(function.get("cases", [])):
        for text in as_list(case.get("preconditions")):
            preconditions.setdefault(str(compact_case_text(text, "precondition")), []).append(idx)

        for input_name, raw_value in (case.get("inputs") or {}).items():
            inputs.setdefault(str(input_name), {})
            for value in as_list(raw_value):
                normalized = normalize_scalar(value)
                key = json.dumps(normalized, ensure_ascii=False, default=str)
                inputs[str(input_name)].setdefault(key, {"display": normalized, "cases": []})
                inputs[str(input_name)][key]["cases"].append(idx)

        for text in as_list(case.get("returns")):
            returns.setdefault(str(compact_case_text(text, "return")), []).append(idx)
        for text in as_list(case.get("exceptions")):
            exceptions.setdefault(str(compact_case_text(text, "exception")), []).append(idx)
        for text in as_list(case.get("logs")):
            logs.setdefault(str(compact_case_text(text, "log")), []).append(idx)

    return preconditions, inputs, returns, exceptions, logs


def clear_case_columns(ws: Worksheet, layout: FunctionSheetLayout) -> None:
    for col in range(layout.case_start_col, layout.last_case_col + 1):
        ws.cell(layout.utc_row, col).value = None
        for row in range(layout.condition_header_row, ws.max_row + 1):
            ws.cell(row, col).value = None


def rebuild_function_body(
    ws: Worksheet,
    template_ws: Worksheet,
    layout: FunctionSheetLayout,
    function: Dict[str, Any],
) -> None:
    cases = function["cases"]
    preconditions, inputs, returns, exceptions, logs = grouped_case_rows(function)

    if ws.max_row >= 11:
        ws.delete_rows(11, ws.max_row - 10)

    clear_case_columns(ws, layout)

    for index in range(layout.case_start_col, layout.last_case_col + 1):
        ws.cell(layout.utc_row, index).value = None

    for case_index, case in enumerate(cases):
        if case_index + layout.case_start_col > layout.last_case_col:
            raise SystemExit(
                f"Template supports only {layout.last_case_col - layout.case_start_col + 1} UTC columns, "
                f"but {function['code']} needs {len(cases)} cases.",
            )
        ws.cell(layout.utc_row, layout.case_start_col + case_index).value = case["utcId"]

    write_value(ws, "C2", function["code"])
    write_value(ws, "L2", function["function_name"])
    write_value(ws, "C3", function["created_by"])
    write_value(ws, "L3", function["executed_by"])
    write_value(ws, "C4", function["loc"])
    write_value(ws, "L4", function["lack_of_test_cases"])
    write_value(ws, "C5", compact_test_requirement(function["test_requirement"]), wrap_text=True)

    row = 11
    for text, case_indexes in preconditions.items():
        copy_row_style(template_ws, 11, ws, row)
        write_content(ws, row, layout, text, wrap_text=True)
        write_case_mark(ws, row, layout, case_indexes)
        row += 1

    for input_name, value_map in inputs.items():
        copy_row_style(template_ws, 13, ws, row)
        copy_row_merges(template_ws, 13, ws, row)
        write_row_label(ws, row, layout, input_name)
        row += 1
        for entry in value_map.values():
            copy_row_style(template_ws, 14, ws, row)
            write_content(ws, row, layout, entry["display"], wrap_text=True)
            write_case_mark(ws, row, layout, entry["cases"])
            row += 1

    confirm_header_row = row
    copy_row_style(template_ws, 15, ws, confirm_header_row)
    copy_row_merges(template_ws, 15, ws, confirm_header_row)
    ws.cell(confirm_header_row, 1).value = "Confirm"
    ws.cell(confirm_header_row, 2).value = "Return"
    apply_content_font(ws.cell(confirm_header_row, 1))
    apply_content_font(ws.cell(confirm_header_row, 2))
    row += 1

    for text, case_indexes in returns.items():
        copy_row_style(template_ws, 16, ws, row)
        write_content(ws, row, layout, text, wrap_text=True)
        write_case_mark(ws, row, layout, case_indexes)
        row += 1

    if exceptions:
        copy_row_style(template_ws, 18, ws, row)
        copy_row_merges(template_ws, 18, ws, row)
        write_row_label(ws, row, layout, "Exception")
        row += 1
        for text, case_indexes in exceptions.items():
            copy_row_style(template_ws, 19, ws, row)
            write_content(ws, row, layout, text, wrap_text=True)
            write_case_mark(ws, row, layout, case_indexes)
            row += 1

    if logs:
        copy_row_style(template_ws, 18, ws, row)
        copy_row_merges(template_ws, 18, ws, row)
        write_row_label(ws, row, layout, "Log message")
        row += 1
        for text, case_indexes in logs.items():
            copy_row_style(template_ws, 19, ws, row)
            write_content(ws, row, layout, text, wrap_text=True)
            write_case_mark(ws, row, layout, case_indexes)
            row += 1

    result_row = row
    copy_row_style(template_ws, 22, ws, result_row)
    copy_row_merges(template_ws, 22, ws, result_row)
    ws.cell(result_row, 1).value = "Result"
    ws.cell(result_row, 2).value = "Type(N : Normal, A : Abnormal, B : Boundary)"
    apply_content_font(ws.cell(result_row, 1))
    apply_content_font(ws.cell(result_row, 2))

    status_row = result_row + 1
    copy_row_style(template_ws, 23, ws, status_row)
    copy_row_merges(template_ws, 23, ws, status_row)
    ws.cell(status_row, 2).value = "Passed/Failed"
    apply_content_font(ws.cell(status_row, 2))

    executed_row = result_row + 2
    copy_row_style(template_ws, 24, ws, executed_row)
    copy_row_merges(template_ws, 24, ws, executed_row)
    ws.cell(executed_row, 2).value = "Executed Date"
    apply_content_font(ws.cell(executed_row, 2))

    defect_row = result_row + 3
    copy_row_style(template_ws, 25, ws, defect_row)
    copy_row_merges(template_ws, 25, ws, defect_row)
    ws.cell(defect_row, 2).value = "Defect ID"
    apply_content_font(ws.cell(defect_row, 2))

    for index, case in enumerate(cases):
        col = layout.case_start_col + index
        ws.cell(result_row, col).value = case.get("type") or ""
        ws.cell(status_row, col).value = case.get("status") or "U"
        ws.cell(executed_row, col).value = case.get("executedDate") or ""
        ws.cell(defect_row, col).value = case.get("defectId") or ""
        for row_index in (result_row, status_row, executed_row, defect_row):
            apply_content_font(ws.cell(row_index, col))

        executed_alignment = copy(ws.cell(executed_row, col).alignment)
        executed_alignment.textRotation = 90
        executed_alignment.horizontal = "center"
        executed_alignment.vertical = "center"
        executed_alignment.wrap_text = False
        ws.cell(executed_row, col).alignment = executed_alignment

    ws.row_dimensions[executed_row].height = max(ws.row_dimensions[executed_row].height or 0, 42)

    status_range = (
        f"{openpyxl.utils.get_column_letter(layout.case_start_col)}{status_row}:"
        f"{openpyxl.utils.get_column_letter(layout.case_start_col + len(cases) - 1)}{status_row}"
    )
    type_range = (
        f"{openpyxl.utils.get_column_letter(layout.case_start_col)}{result_row}:"
        f"{openpyxl.utils.get_column_letter(layout.case_start_col + len(cases) - 1)}{result_row}"
    )
    total_range = (
        f"{openpyxl.utils.get_column_letter(layout.case_start_col)}{layout.utc_row}:"
        f"{openpyxl.utils.get_column_letter(layout.case_start_col + len(cases) - 1)}{layout.utc_row}"
    )

    write_value(ws, layout.passed_cell, f'=COUNTIF({status_range},"P")')
    write_value(ws, layout.failed_cell, f'=COUNTIF({status_range},"F")')
    write_value(ws, layout.untested_cell, f'=COUNTIF({status_range},"U")')
    write_value(ws, layout.n_count_cell, f'=COUNTIF({type_range},"N")')
    write_value(ws, layout.a_count_cell, f'=COUNTIF({type_range},"A")')
    write_value(ws, layout.b_count_cell, f'=COUNTIF({type_range},"B")')
    write_value(ws, layout.total_cell, f'=COUNTA({total_range})')
    last_row = max(defect_row, 25)
    ws.print_area = f"A1:T{last_row}"
    trim_function_sheet(ws, last_row)


def sanitize_sheet_name(name: str) -> str:
    clean = re.sub(r'[:\\/\?\*\[\]]', "_", name.strip())
    return clean[:31] or "Function"


def ensure_unique_name(existing: Iterable[str], desired: str) -> str:
    existing_names = set(existing)
    if desired not in existing_names:
        return desired
    base = desired[:28]
    counter = 2
    while True:
        candidate = f"{base}-{counter}"
        if candidate not in existing_names:
            return candidate
        counter += 1


def populate_cover(ws: Worksheet, meta: Dict[str, Any], workbook_spec: Dict[str, Any]) -> None:
    write_value(ws, "B4", meta["project_name"])
    write_value(ws, "B5", meta["project_code"])
    write_value(ws, "B6", workbook_spec["document_code"])
    write_value(ws, "F4", meta["creator"])
    write_value(ws, "F5", meta["reviewer"])
    write_value(ws, "F6", meta["issue_date"])
    write_value(ws, "F7", workbook_spec["version"])

    if ws.max_row >= 12:
        for row in range(12, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).value = None

    change_row = 11
    for entry in workbook_spec.get("change_log", []):
        values = [
            entry.get("effective_date") or "",
            entry.get("version") or "",
            entry.get("change_item") or "",
            entry.get("mode") or "A",
            entry.get("change_description") or "",
            entry.get("reference") or "",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(change_row, col)
            cell.value = value
            apply_content_font(cell)
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            cell.alignment = alignment
        change_row += 1


def build_function_list_rows(functions: List[Dict[str, Any]]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for index, function in enumerate(functions, start=1):
        rows.append(
            [
                index,
                function["requirement_name"],
                function["class_name"],
                function["function_name"],
                function["code"],
                function["sheet_name"],
                function["description"],
                function["precondition_summary"],
            ]
        )
    return rows


def populate_function_list(ws: Worksheet, template_ws: Worksheet, meta: Dict[str, Any], functions: List[Dict[str, Any]]) -> None:
    write_value(ws, "E4", meta["project_name"])
    write_value(ws, "E5", meta["project_code"])
    write_value(ws, "E6", template_ws["E6"].value)
    write_value(ws, "E7", meta["environment_description"], wrap_text=True)

    if ws.max_row >= 11:
        ws.delete_rows(11, ws.max_row - 10)

    template_row = 11
    for offset, values in enumerate(build_function_list_rows(functions), start=11):
        copy_row_style(template_ws, template_row, ws, offset)
        for col, value in enumerate(values, start=1):
            cell = ws.cell(offset, col)
            cell.value = value
            apply_content_font(cell)
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            cell.alignment = alignment
    ws.print_area = f"'Function List'!$A$1:$H${max(38, 10 + len(functions))}"


def escape_sheet_ref(sheet_name: str) -> str:
    return sheet_name.replace("'", "''")


def populate_test_report(ws: Worksheet, template_ws: Worksheet, meta: Dict[str, Any], functions: List[Dict[str, Any]]) -> None:
    ws["B4"] = meta["project_name"]
    ws["B5"] = meta["project_code"]
    ws["B6"] = template_ws["B6"].value
    ws["F4"] = meta["creator"]
    ws["F5"] = meta["reviewer"]
    ws["F6"] = meta["issue_date"]
    for ref in ["B4", "B5", "B6", "F4", "F5", "F6"]:
        apply_content_font(ws[ref])

    ws["B7"] = meta["report_notes"]
    apply_content_font(ws["B7"])
    ws["B7"].alignment = Alignment(
        horizontal=ws["B7"].alignment.horizontal,
        vertical=ws["B7"].alignment.vertical,
        wrap_text=True,
    )

    if ws.max_row >= 12:
        ws.delete_rows(12, ws.max_row - 11)

    data_start_row = 12
    for index, function in enumerate(functions, start=data_start_row):
        copy_row_style(template_ws, 12, ws, index)
        sheet_ref = escape_sheet_ref(function["sheet_name"])
        values = [
            index - data_start_row + 1,
            function["code"],
            f"='{sheet_ref}'!$A$7",
            f"='{sheet_ref}'!$C$7",
            f"='{sheet_ref}'!$F$7",
            f"='{sheet_ref}'!$L$7",
            f"='{sheet_ref}'!$M$7",
            f"='{sheet_ref}'!$N$7",
            f"='{sheet_ref}'!$O$7",
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(index, col)
            cell.value = value
            apply_content_font(cell)

    blank_row = data_start_row + len(functions)
    copy_row_style(template_ws, 95, ws, blank_row)

    subtotal_row = blank_row + 1
    copy_row_style(template_ws, 96, ws, subtotal_row)
    ws.cell(subtotal_row, 2).value = "Sub total"
    apply_content_font(ws.cell(subtotal_row, 2))
    for col in range(3, 10):
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.cell(subtotal_row, col).value = f"=SUM({col_letter}{data_start_row}:{col_letter}{blank_row - 1})"
        apply_content_font(ws.cell(subtotal_row, col))

    second_blank_row = subtotal_row + 1
    copy_row_style(template_ws, 97, ws, second_blank_row)

    metrics = [
        ("Test coverage", f'=IF(I{subtotal_row}=0,0,(C{subtotal_row}+D{subtotal_row})*100/I{subtotal_row})'),
        ("Test successful coverage", f'=IF(I{subtotal_row}=0,0,C{subtotal_row}*100/I{subtotal_row})'),
        ("Normal case", f'=IF(I{subtotal_row}=0,0,F{subtotal_row}*100/I{subtotal_row})'),
        ("Abnormal case", f'=IF(I{subtotal_row}=0,0,G{subtotal_row}*100/I{subtotal_row})'),
        ("Boundary case", f'=IF(I{subtotal_row}=0,0,H{subtotal_row}*100/I{subtotal_row})'),
    ]
    metric_template_rows = [98, 99, 100, 101, 102]
    for index, (label, formula) in enumerate(metrics):
        target_row = second_blank_row + 1 + index
        copy_row_style(template_ws, metric_template_rows[index], ws, target_row)
        ws.cell(target_row, 2).value = label
        ws.cell(target_row, 4).value = formula
        ws.cell(target_row, 5).value = "%"
        apply_content_font(ws.cell(target_row, 2))
        apply_content_font(ws.cell(target_row, 4))
        apply_content_font(ws.cell(target_row, 5))
    ws.print_area = f"'Test Report'!$A$1:$I${max(120, second_blank_row + 5)}"


def build_workbook(template_path: Path, spec_path: Path) -> Path:
    spec = read_json(spec_path)
    meta = spec["meta"]
    workbook_spec = spec["workbook"]

    wb = openpyxl.load_workbook(template_path)
    donor_wb = openpyxl.load_workbook(template_path)
    template_sheet_name = pick_template_sheet_name(wb)
    template_sheet = wb[template_sheet_name]
    donor_template_sheet = donor_wb[template_sheet_name]
    layout = detect_layout(template_sheet)

    for name in list(wb.sheetnames):
        if name in HELPER_SHEETS or name == template_sheet_name:
            continue
        wb.remove(wb[name])

    function_list_template = donor_wb["Function List"]
    test_report_template = donor_wb["Test Report"]
    functions = workbook_spec["functions"]

    for function in functions:
        ws = wb.copy_worksheet(template_sheet)
        copy_sheet_view(donor_template_sheet, ws)
        clear_sheet_artifacts(ws)
        final_name = ensure_unique_name(wb.sheetnames, sanitize_sheet_name(function["sheet_name"]))
        ws.title = final_name
        function["sheet_name"] = final_name
        function["created_by"] = meta["creator"]
        function["executed_by"] = meta["executed_by"]
        rebuild_function_body(ws, donor_template_sheet, layout, function)

    wb.remove(template_sheet)

    populate_cover(wb["Cover"], meta, workbook_spec)
    populate_function_list(wb["Function List"], function_list_template, meta, functions)
    populate_test_report(wb["Test Report"], test_report_template, meta, functions)

    for name in wb.sheetnames:
        clear_sheet_artifacts(wb[name])

    output_path = Path(workbook_spec["output_file"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    openpyxl.load_workbook(output_path)
    return output_path.resolve()


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Generate a single unit-test workbook from executed evidence.")
    parser.add_argument("--spec", required=True, help="Path to the workbook spec JSON.")
    parser.add_argument("--template", required=True, help="Path to the Unit Test workbook template.")
    args = parser.parse_args(argv)

    output_path = build_workbook(Path(args.template), Path(args.spec))
    sys.stdout.buffer.write(f"{output_path}\n".encode("utf-8", errors="backslashreplace"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
