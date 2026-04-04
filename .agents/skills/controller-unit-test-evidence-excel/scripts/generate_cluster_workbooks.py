#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from copy import copy
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.cell.cell import MergedCell

HELPER_SHEETS = {"Guideline", "Cover", "Function List", "Test Report", "Statistics"}


@dataclass(frozen=True)
class FunctionSheetLayout:
    case_start_col: int
    utc_row: int
    condition_row: int
    confirm_row: int
    result_row: int
    passed_cell: str
    failed_cell: str
    untested_cell: str
    n_count_cell: str
    a_count_cell: str
    b_count_cell: str
    total_cell: str


def read_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def find_cell(ws: openpyxl.worksheet.worksheet.Worksheet, needle: str) -> Optional[Tuple[int, int]]:
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value == needle:
                return cell.row, cell.column
    return None


def find_row_in_col_a(ws: openpyxl.worksheet.worksheet.Worksheet, needle: str) -> Optional[int]:
    for row in range(1, ws.max_row + 1):
        if ws.cell(row, 1).value == needle:
            return row
    return None


def utc_info(
    ws: openpyxl.worksheet.worksheet.Worksheet,
) -> Optional[Tuple[int, int, int]]:
    pos = find_cell(ws, "UTCID01")
    if not pos:
        return None
    utc_row, utc_col = pos
    last_col = utc_col
    for col in range(utc_col, ws.max_column + 1):
        value = ws.cell(utc_row, col).value
        if isinstance(value, str) and value.startswith("UTCID"):
            last_col = col
    return utc_row, utc_col, last_col


def pick_template_sheet_name(wb: openpyxl.Workbook) -> str:
    best_name: Optional[str] = None
    best_score = (-10**9, -10**9)

    for name in wb.sheetnames:
        if name in HELPER_SHEETS:
            continue
        ws = wb[name]
        utc = utc_info(ws)
        if not utc:
            continue
        if not find_row_in_col_a(ws, "Condition") or not find_row_in_col_a(ws, "Confirm") or not find_row_in_col_a(ws, "Result"):
            continue
        utc_row, utc_col, last_col = utc
        utc_count = 0
        for col in range(utc_col, last_col + 1):
            value = ws.cell(utc_row, col).value
            if isinstance(value, str) and value.startswith("UTCID"):
                utc_count += 1
        score = (-ws.max_row, utc_count)
        if score > best_score:
            best_name = name
            best_score = score

    if not best_name:
        raise SystemExit("Could not find a function-sheet template in the workbook.")
    return best_name


def cell_below_label(ws: openpyxl.worksheet.worksheet.Worksheet, label: str) -> str:
    pos = find_cell(ws, label)
    if not pos:
        raise SystemExit(f"Template sheet is missing label: {label}")
    row, col = pos
    return openpyxl.utils.get_column_letter(col) + str(row + 1)


def detect_layout(
    wb: openpyxl.Workbook,
    template_sheet_name: str,
) -> FunctionSheetLayout:
    ws = wb[template_sheet_name]
    utc = utc_info(ws)
    if not utc:
        raise SystemExit("Template function sheet is missing UTC columns.")
    utc_row, case_start_col, _ = utc
    condition_row = find_row_in_col_a(ws, "Condition")
    confirm_row = find_row_in_col_a(ws, "Confirm")
    result_row = find_row_in_col_a(ws, "Result")
    n_a_b_pos = find_cell(ws, "N/A/B")
    total_pos = find_cell(ws, "Total Test Cases")
    if not condition_row or not confirm_row or not result_row or not n_a_b_pos or not total_pos:
        raise SystemExit("Template function sheet is missing core UTC rows.")

    return FunctionSheetLayout(
        case_start_col=case_start_col,
        utc_row=utc_row,
        condition_row=condition_row,
        confirm_row=confirm_row,
        result_row=result_row,
        passed_cell=cell_below_label(ws, "Passed"),
        failed_cell=cell_below_label(ws, "Failed"),
        untested_cell=cell_below_label(ws, "Untested"),
        n_count_cell=openpyxl.utils.get_column_letter(n_a_b_pos[1]) + str(n_a_b_pos[0] + 1),
        a_count_cell=openpyxl.utils.get_column_letter(n_a_b_pos[1] + 1) + str(n_a_b_pos[0] + 1),
        b_count_cell=openpyxl.utils.get_column_letter(n_a_b_pos[1] + 2) + str(n_a_b_pos[0] + 1),
        total_cell=openpyxl.utils.get_column_letter(total_pos[1]) + str(total_pos[0] + 1),
    )


def clear_range(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def copy_row_style(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    source_row: int,
    target_row: int,
) -> None:
    source_dimension = ws.row_dimensions[source_row]
    target_dimension = ws.row_dimensions[target_row]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed

    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        if isinstance(source, MergedCell):
            continue
        target = ws.cell(target_row, col)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)
        target.fill = copy(source.fill)
        target.font = copy(source.font)
        target.border = copy(source.border)


def expand_function_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    layout: FunctionSheetLayout,
    condition_rows_needed: int,
    confirm_rows_needed: int,
) -> FunctionSheetLayout:
    adjusted = layout
    condition_capacity = adjusted.confirm_row - adjusted.condition_row - 1
    extra_condition_rows = max(condition_rows_needed - condition_capacity, 0)
    if extra_condition_rows > 0:
        ws.insert_rows(adjusted.confirm_row, extra_condition_rows)
        for offset in range(extra_condition_rows):
            copy_row_style(ws, adjusted.condition_row + 1, adjusted.confirm_row + offset)
        adjusted = replace(
            adjusted,
            confirm_row=adjusted.confirm_row + extra_condition_rows,
            result_row=adjusted.result_row + extra_condition_rows,
        )

    confirm_capacity = adjusted.result_row - adjusted.confirm_row - 1
    extra_confirm_rows = max(confirm_rows_needed - confirm_capacity, 0)
    if extra_confirm_rows > 0:
        ws.insert_rows(adjusted.result_row, extra_confirm_rows)
        for offset in range(extra_confirm_rows):
            copy_row_style(ws, adjusted.confirm_row + 1, adjusted.result_row + offset)
        adjusted = replace(adjusted, result_row=adjusted.result_row + extra_confirm_rows)

    return adjusted


def unmerge_content_rows(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
) -> None:
    target_ranges = [
        merged_range
        for merged_range in list(ws.merged_cells.ranges)
        if (
            start_row <= merged_range.min_row <= merged_range.max_row <= end_row
            and 2 <= merged_range.min_col <= merged_range.max_col <= 4
        )
    ]

    for merged_range in target_ranges:
        top_left = ws.cell(merged_range.min_row, merged_range.min_col)
        preserved_value = top_left.value
        preserved_style = copy(top_left._style)
        preserved_number_format = top_left.number_format
        preserved_protection = copy(top_left.protection)
        preserved_alignment = copy(top_left.alignment)
        preserved_fill = copy(top_left.fill)
        preserved_font = copy(top_left.font)
        preserved_border = copy(top_left.border)

        ws.unmerge_cells(str(merged_range))
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                cell = ws.cell(row, col)
                cell._style = copy(preserved_style)
                cell.number_format = preserved_number_format
                cell.protection = copy(preserved_protection)
                cell.alignment = copy(preserved_alignment)
                cell.fill = copy(preserved_fill)
                cell.font = copy(preserved_font)
                cell.border = copy(preserved_border)
        ws.cell(merged_range.min_row, merged_range.min_col).value = preserved_value


def purge_cells(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws._cells.pop((row, col), None)


def apply_content_font(cell: openpyxl.cell.cell.Cell) -> None:
    font = copy(cell.font)
    font.name = "Calibri"
    font.size = 12
    cell.font = font


def write_value(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    col: int,
    value: Any,
    content_font: bool = True,
) -> None:
    cell = ws.cell(row, col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                cell = ws.cell(merged_range.min_row, merged_range.min_col)
                break
        else:
            return
    cell.value = value
    if content_font:
        apply_content_font(cell)


def set_header_value(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    label: str,
    offset_cols: int,
    value: Any,
) -> None:
    pos = find_cell(ws, label)
    if not pos:
        return
    row, col = pos
    write_value(ws, row, col + offset_cols, value)


def flatten_assertions(results_payload: Dict[str, Any]) -> List[Dict[str, str]]:
    flattened: List[Dict[str, str]] = []
    for suite in results_payload.get("testResults", []):
        for assertion in suite.get("assertionResults", []):
            flattened.append(
                {
                    "status": assertion.get("status", ""),
                    "title": assertion.get("title", ""),
                    "fullName": assertion.get("fullName", ""),
                }
            )
    return flattened


def apply_evidence(functions: List[Dict[str, Any]], assertions: List[Dict[str, str]], executed_date: str) -> None:
    for function in functions:
        for case in function.get("cases", []):
            test_key = str(case.get("test_key") or "").strip()
            matches = [
                assertion
                for assertion in assertions
                if assertion["title"].startswith(test_key) or assertion["fullName"].startswith(test_key)
            ]
            if not matches:
                case["status"] = "U"
                case["executed_date"] = ""
                continue

            statuses = {match["status"] for match in matches}
            case["status"] = "F" if "failed" in statuses else "P"
            case["executed_date"] = executed_date


def write_marks(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    row: int,
    start_col: int,
    case_indexes: Sequence[int],
) -> None:
    for idx in case_indexes:
        write_value(ws, row, start_col + idx, "O")


def normalize_scalar(value: Any) -> Any:
    if value is None:
        return "null"
    return value


def as_list(value: Any) -> List[Any]:
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def group_case_rows(function: Dict[str, Any]) -> Tuple[Dict[str, List[int]], Dict[str, Dict[str, Dict[str, Any]]], Dict[str, List[int]], Dict[str, List[int]]]:
    preconditions: Dict[str, List[int]] = {}
    inputs: Dict[str, Dict[str, Dict[str, Any]]] = {}
    returns: Dict[str, List[int]] = {}
    exceptions: Dict[str, List[int]] = {}

    for idx, case in enumerate(function.get("cases", [])):
        for text in as_list(case.get("preconditions")):
            preconditions.setdefault(str(text), []).append(idx)

        for input_name, raw_value in (case.get("inputs") or {}).items():
            inputs.setdefault(str(input_name), {})
            for value in as_list(raw_value):
                normalized = normalize_scalar(value)
                key = json.dumps(normalized, ensure_ascii=False, default=str)
                inputs[str(input_name)].setdefault(
                    key,
                    {"display": normalized, "cases": []},
                )
                inputs[str(input_name)][key]["cases"].append(idx)

        for text in as_list(case.get("returns")):
            returns.setdefault(str(text), []).append(idx)
        for text in as_list(case.get("exceptions")):
            exceptions.setdefault(str(text), []).append(idx)

    return preconditions, inputs, returns, exceptions


def populate_function_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    layout: FunctionSheetLayout,
    function: Dict[str, Any],
) -> None:
    cases = function.get("cases", [])
    preconditions, inputs, returns, exceptions = group_case_rows(function)
    condition_rows_needed = len(preconditions) + sum(1 + len(value_map) for value_map in inputs.values())
    confirm_rows_needed = len(returns) + (1 + len(exceptions) if exceptions else 0)
    adjusted_layout = expand_function_sheet(ws, layout, condition_rows_needed, confirm_rows_needed)
    unmerge_content_rows(ws, adjusted_layout.condition_row + 1, adjusted_layout.result_row + 3)

    clear_range(ws, adjusted_layout.utc_row, adjusted_layout.result_row + 3, 1, ws.max_column)

    write_value(ws, adjusted_layout.condition_row, 1, "Condition", content_font=False)
    write_value(ws, adjusted_layout.condition_row, 2, "Precondition ", content_font=False)
    write_value(ws, adjusted_layout.confirm_row, 1, "Confirm", content_font=False)
    write_value(ws, adjusted_layout.confirm_row, 2, "Return", content_font=False)
    write_value(ws, adjusted_layout.result_row, 1, "Result", content_font=False)
    write_value(
        ws,
        adjusted_layout.result_row,
        2,
        "Type(N : Normal, A : Abnormal, B : Boundary)",
        content_font=False,
    )
    write_value(ws, adjusted_layout.result_row + 1, 2, "Passed/Failed", content_font=False)
    write_value(ws, adjusted_layout.result_row + 2, 2, "Executed Date", content_font=False)
    write_value(ws, adjusted_layout.result_row + 3, 2, "Defect ID", content_font=False)

    set_header_value(ws, "Function Code", 2, function.get("function_code") or "")
    set_header_value(ws, "Function Name", 6, function.get("function_name") or "")
    set_header_value(ws, "Created By", 2, function.get("created_by") or "")
    set_header_value(ws, "Executed By", 6, function.get("executed_by") or function.get("created_by") or "")
    set_header_value(ws, "Lines  of code", 2, function.get("loc") or "")
    set_header_value(ws, "Test requirement", 2, function.get("test_requirement") or "")

    for idx in range(len(cases)):
        write_value(ws, adjusted_layout.utc_row, adjusted_layout.case_start_col + idx, f"UTCID{idx + 1:02d}")

    row = adjusted_layout.condition_row + 1
    for text, case_indexes in preconditions.items():
        write_value(ws, row, 4, text)
        write_marks(ws, row, adjusted_layout.case_start_col, case_indexes)
        row += 1

    for input_name, value_map in inputs.items():
        write_value(ws, row, 2, input_name)
        row += 1
        for entry in value_map.values():
            write_value(ws, row, 4, entry["display"])
            write_marks(ws, row, adjusted_layout.case_start_col, entry["cases"])
            row += 1

    row = adjusted_layout.confirm_row + 1
    for text, case_indexes in returns.items():
        write_value(ws, row, 4, text)
        write_marks(ws, row, adjusted_layout.case_start_col, case_indexes)
        row += 1

    if exceptions:
        write_value(ws, row, 2, "Exception")
        row += 1
        for text, case_indexes in exceptions.items():
            write_value(ws, row, 4, text)
            write_marks(ws, row, adjusted_layout.case_start_col, case_indexes)
            row += 1

    passed_count = 0
    failed_count = 0
    n_count = 0
    a_count = 0
    b_count = 0

    for idx, case in enumerate(cases):
        case_type = str(case.get("type") or "").upper()
        case_status = str(case.get("status") or "").upper()
        write_value(ws, adjusted_layout.result_row, adjusted_layout.case_start_col + idx, case_type)
        write_value(ws, adjusted_layout.result_row + 1, adjusted_layout.case_start_col + idx, case_status)
        write_value(
            ws,
            adjusted_layout.result_row + 2,
            adjusted_layout.case_start_col + idx,
            case.get("executed_date") or "",
        )
        write_value(
            ws,
            adjusted_layout.result_row + 3,
            adjusted_layout.case_start_col + idx,
            case.get("defect_id") or "",
        )

        if case_status == "P":
            passed_count += 1
        elif case_status == "F":
            failed_count += 1

        if case_type == "N":
            n_count += 1
        elif case_type == "A":
            a_count += 1
        elif case_type == "B":
            b_count += 1

    ws[layout.passed_cell] = passed_count
    ws[layout.failed_cell] = failed_count
    ws[layout.untested_cell] = max(len(cases) - passed_count - failed_count, 0)
    ws[layout.n_count_cell] = n_count
    ws[layout.a_count_cell] = a_count
    ws[layout.b_count_cell] = b_count
    ws[layout.total_cell] = len(cases)

    for ref in [
        layout.passed_cell,
        layout.failed_cell,
        layout.untested_cell,
        layout.n_count_cell,
        layout.a_count_cell,
        layout.b_count_cell,
        layout.total_cell,
    ]:
        apply_content_font(ws[ref])


def update_cover(
    wb: openpyxl.Workbook,
    meta: Dict[str, Any],
    workbook_spec: Dict[str, Any],
) -> None:
    if "Cover" not in wb.sheetnames:
        return
    ws = wb["Cover"]
    write_value(ws, 4, 2, meta.get("project_name") or "")
    write_value(ws, 5, 2, meta.get("project_code") or "")
    write_value(ws, 4, 6, meta.get("creator") or "")
    write_value(ws, 5, 6, meta.get("reviewer") or "")
    write_value(ws, 6, 2, f'=B5&"_"&"Test Report"&"_"&"v{workbook_spec["version"]}"')
    write_value(ws, 6, 6, workbook_spec.get("issue_date") or meta.get("issue_date") or "")
    write_value(ws, 7, 6, workbook_spec["version"])

    clear_range(ws, 24, max(ws.max_row, 40), 1, 6)
    row = 24
    for entry in workbook_spec.get("change_log", []):
        values = [
            entry.get("effective_date") or "",
            entry.get("version") or "",
            entry.get("change_item") or "",
            entry.get("mode") or "A",
            entry.get("change_description") or "",
            entry.get("reference") or "",
        ]
        for col, value in enumerate(values, start=1):
            write_value(ws, row, col, value)
        row += 1


def update_function_list(
    wb: openpyxl.Workbook,
    rows: List[Dict[str, Any]],
) -> None:
    if "Function List" not in wb.sheetnames:
        return
    ws = wb["Function List"]
    clear_range(ws, 11, max(ws.max_row, 120), 1, 7)
    for row_idx, row_data in enumerate(rows, start=11):
        values = [
            row_data.get("no"),
            row_data.get("class_name"),
            row_data.get("function_name"),
            row_data.get("function_code"),
            row_data.get("sheet_name"),
            row_data.get("description"),
            row_data.get("precondition"),
        ]
        for col, value in enumerate(values, start=1):
            write_value(ws, row_idx, col, value)
    purge_start = 11 + len(rows)
    if purge_start <= max(ws.max_row, 120):
        purge_cells(ws, purge_start, max(ws.max_row, 120), 1, 7)


def escape_sheet_ref(sheet_name: str) -> str:
    return sheet_name.replace("'", "''")


def update_test_report(
    wb: openpyxl.Workbook,
    meta: Dict[str, Any],
    workbook_spec: Dict[str, Any],
    layout: FunctionSheetLayout,
) -> None:
    if "Test Report" not in wb.sheetnames:
        return
    ws = wb["Test Report"]
    write_value(ws, 4, 2, meta.get("project_name") or "")
    write_value(ws, 4, 6, meta.get("creator") or "")
    write_value(ws, 5, 2, meta.get("project_code") or "")
    write_value(ws, 5, 6, meta.get("reviewer") or "")
    write_value(ws, 6, 2, f'=B5&"_"&"Test Report"&"_"&"v{workbook_spec["version"]}"')
    write_value(ws, 6, 6, workbook_spec.get("issue_date") or meta.get("issue_date") or "")

    clear_range(ws, 12, max(ws.max_row, 120), 1, 9)
    for index, function in enumerate(workbook_spec.get("functions", []), start=12):
        sheet_ref = escape_sheet_ref(function["sheet_name"])
        values = [
            index - 11,
            function["function_code"],
            f"='{sheet_ref}'!{layout.passed_cell}",
            f"='{sheet_ref}'!{layout.failed_cell}",
            f"='{sheet_ref}'!{layout.untested_cell}",
            f"='{sheet_ref}'!{layout.n_count_cell}",
            f"='{sheet_ref}'!{layout.a_count_cell}",
            f"='{sheet_ref}'!{layout.b_count_cell}",
            f"='{sheet_ref}'!{layout.total_cell}",
        ]
        for col, value in enumerate(values, start=1):
            write_value(ws, index, col, value)
    purge_start = 12 + len(workbook_spec.get("functions", []))
    if purge_start <= max(ws.max_row, 120):
        purge_cells(ws, purge_start, max(ws.max_row, 120), 1, 9)


def sanitize_sheet_name(name: str) -> str:
    value = re.sub(r'[:\\/\?\*\[\]]', "_", name.strip())
    return value[:31] or "Function"


def ensure_unique_name(existing: Iterable[str], desired: str) -> str:
    existing_names = set(existing)
    if desired not in existing_names:
        return desired
    base = desired[:28]
    counter = 2
    while True:
        candidate = f"{base}-{counter}"
        if candidate not in existing_names:
            return candidate
        counter += 1


def build_workbook(
    template_path: Path,
    meta: Dict[str, Any],
    workbook_spec: Dict[str, Any],
) -> Path:
    wb = openpyxl.load_workbook(template_path)
    template_sheet_name = pick_template_sheet_name(wb)
    layout = detect_layout(wb, template_sheet_name)

    for name in list(wb.sheetnames):
        if name in HELPER_SHEETS or name == template_sheet_name:
            continue
        wb.remove(wb[name])

    template_sheet = wb[template_sheet_name]
    created_sheet_names: List[str] = []
    for function in workbook_spec.get("functions", []):
        ws = wb.copy_worksheet(template_sheet)
        sheet_name = ensure_unique_name(wb.sheetnames, sanitize_sheet_name(function["sheet_name"]))
        ws.title = sheet_name
        function["sheet_name"] = sheet_name
        created_sheet_names.append(sheet_name)
        populate_function_sheet(ws, layout, function)

    wb.remove(template_sheet)
    update_cover(wb, meta, workbook_spec)
    update_function_list(wb, workbook_spec.get("function_list", []))
    update_test_report(wb, meta, workbook_spec, layout)

    output_path = Path(workbook_spec["output_file"])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    openpyxl.load_workbook(output_path)
    return output_path.resolve()


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Generate versioned UTC workbooks by module cluster.")
    parser.add_argument("--spec", required=True, help="Path to the workbook spec JSON.")
    parser.add_argument("--results", required=True, help="Path to the Jest JSON result file.")
    parser.add_argument("--template", required=True, help="Path to the template workbook.")
    args = parser.parse_args(argv)

    spec = read_json(Path(args.spec))
    results = read_json(Path(args.results))
    assertions = flatten_assertions(results)
    meta = spec.get("meta") or {}
    executed_date = str(meta.get("executed_date") or meta.get("issue_date") or "")

    output_paths: List[Path] = []
    for workbook_spec in spec.get("workbooks", []):
        apply_evidence(workbook_spec.get("functions", []), assertions, executed_date)
        output_paths.append(build_workbook(Path(args.template), meta, workbook_spec))

    for path in output_paths:
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
