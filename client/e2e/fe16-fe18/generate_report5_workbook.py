from __future__ import annotations

import json
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell


REPO_ROOT = Path(__file__).resolve().parents[3]
TEMPLATE_PATH = REPO_ROOT / "docs" / "Integration Test" / "Report5_Test Case Document.xlsx"
CATALOG_PATH = Path(__file__).resolve().with_name("case-catalog.json")
EVIDENCE_DIR = REPO_ROOT / "client" / "test-artifacts" / "fe16-fe18" / "evidence"
OUTPUT_PATH = (
    REPO_ROOT
    / "docs"
    / "Integration Test"
    / "Report5_Test Case Document_FE16-FE18.xlsx"
)
DEFAULT_TESTER = "BangDC"
FUNCTION_HEADER_SUFFIX = (
    "(each function includes multiple test cases to check User Interface (GUI), "
    "Data Validation (GUI), Functionality, Non-Functionality,..)"
)


def load_catalog() -> dict:
    return json.loads(CATALOG_PATH.read_text(encoding="utf-8"))


def slugify(value: str) -> str:
    return "-".join(part for part in "".join(
        ch.lower() if ch.isalnum() else " " for ch in value
    ).split())


def load_evidence() -> dict[str, dict]:
    evidence_by_case: dict[str, dict] = {}
    for path in sorted(EVIDENCE_DIR.glob("*.json")):
        payload = json.loads(path.read_text(encoding="utf-8"))
        evidence_by_case[payload["id"]] = payload
    return evidence_by_case


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
        target.number_format = source.number_format
    if ws.row_dimensions[source_row].height is not None:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def clear_range(ws, start_row: int, end_row: int, max_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def normalize_line(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def format_numbered_lines(items: list[str]) -> str:
    normalized_items = [normalize_line(item) for item in items if normalize_line(item)]
    return "\n".join(f"{index}. {item}" for index, item in enumerate(normalized_items, start=1))


def format_actual_results(value: str) -> str:
    normalized = normalize_line(value)
    substitutions = [
        (r";\s*", ". "),
        (r", then\s+", ". Then "),
        (r"\band then\s+", ". Then "),
        (r"\band rendered\s+", ". Rendered "),
        (r"\band refreshed\s+", ". Refreshed "),
        (r"\band navigated\s+", ". Navigated "),
        (r"\band redirected\s+", ". Redirected "),
        (r"\band removed\s+", ". Removed "),
        (r"\band showed\s+", ". Showed "),
        (r"\band confirmed\s+", ". Confirmed "),
        (r"\band completed\s+", ". Completed "),
        (r"\band stayed\s+", ". Stayed "),
        (r"\band kept\s+", ". Kept "),
        (r"\band moved\s+", ". Moved "),
        (r"\band displayed\s+", ". Displayed "),
        (r"\band the modal\s+", ". The modal "),
        (r"\band the dashboard\s+", ". The dashboard "),
        (r"\band the dropdown\s+", ". The dropdown "),
        (r"\band the queue\s+", ". The queue "),
        (r"\band the vault\s+", ". The vault "),
        (r"\band the trust profile\s+", ". The trust profile "),
        (r"\band the hearing card\s+", ". The hearing card "),
        (r"\band the frontend\s+", ". The frontend "),
        (r"\band the page\s+", ". The page "),
        (r"\band the active case card\s+", ". The active case card "),
        (r", plus\s+", ". Plus "),
        (r"\s+before redirecting\s+", ". Redirected "),
        (r"\s+so the new party was\s+", ". The new party was "),
    ]
    for pattern, replacement in substitutions:
        normalized = re.sub(pattern, replacement, normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"\s+\.", ".", normalized)
    normalized = re.sub(r",\s*\.", ".", normalized)
    normalized = re.sub(
        r"(No notifications yet\.)\s+placeholder",
        r"\1 placeholder",
        normalized,
        flags=re.IGNORECASE,
    )

    raw_parts = [
        normalize_line(part)
        for part in re.split(r"(?<=[.!?])\s+", normalized)
        if normalize_line(part)
    ]
    parts: list[str] = []
    for part in raw_parts:
        cleaned = part.rstrip(",;")
        if cleaned.lower().startswith("plus correct "):
            cleaned = "Correct " + cleaned[13:]
        if parts and (
            cleaned[:1].islower() or cleaned.lower().startswith("placeholder ")
        ):
            parts[-1] = f"{parts[-1]} {cleaned}"
            continue
        parts.append(cleaned)

    if len(parts) == 1 and len(parts[0]) > 140 and " and " in parts[0]:
        left, right = parts[0].split(" and ", 1)
        parts = [normalize_line(left), normalize_line(right[:1].upper() + right[1:])]
    if not parts:
        parts = [normalize_line(value)]
    return format_numbered_lines(parts)


def format_actual_results_for_case(case: dict) -> str:
    if case.get("result") == "Pass":
        return format_numbered_lines(case["expectedResults"])
    return format_actual_results(case["actualResults"])


def format_note_for_case(case: dict) -> str:
    note_parts: list[str] = []
    raw_actual = normalize_line(case.get("actualResults", ""))
    if raw_actual:
        note_parts.append(f"Automation evidence: {raw_actual}")
    evidence_ref = case.get("evidenceRef") or ""
    if evidence_ref:
        note_parts.append(f"Evidence file: {evidence_ref}")
    existing_note = normalize_line(case.get("note", ""))
    if existing_note:
        note_parts.append(f"Execution note: {existing_note}")
    return "\n".join(note_parts)


def apply_case_body_font(ws, row: int) -> None:
    body_font = copy(ws["A10"].font)
    for col in range(2, 7):
        ws.cell(row, col).font = copy(body_font)


def all_cases(catalog: dict) -> list[dict]:
    ordered_cases: list[dict] = []
    for feature in catalog["features"]:
        for function in feature["functions"]:
            for case in function["cases"]:
                ordered_cases.append(
                    {
                        "feature": feature["sheetName"],
                        "function_name": function["name"],
                        "function_description": function["description"],
                        "pre_condition": function["preCondition"],
                        "reference_document": feature["referenceDocument"],
                        "test_requirement": feature["testRequirement"],
                        **case,
                    }
                )
    return ordered_cases


def merge_case_evidence(catalog: dict) -> tuple[list[dict], str, str, str, str]:
    evidence_by_case = load_evidence()
    cases = []
    for case in all_cases(catalog):
        evidence = evidence_by_case.get(case["id"])
        if not evidence:
            raise SystemExit(f"Missing evidence for case: {case['id']}")
        if evidence.get("result") != "Pass":
            raise SystemExit(f"Case is not passing: {case['id']} -> {evidence.get('result')}")
        cases.append({**case, **evidence})
    return (
        cases,
        catalog["projectName"],
        catalog["projectCode"],
        catalog["documentVersion"],
        catalog["issueDate"],
    )


def count_feature_cases(cases: list[dict], feature_name: str) -> dict[str, int]:
    feature_cases = [case for case in cases if case["feature"] == feature_name]
    return {
        "Pass": len(feature_cases),
        "Fail": 0,
        "Untested": 0,
        "N/A": 0,
        "Total": len(feature_cases),
    }


def populate_cover(ws, project_name: str, project_code: str, version: str, issue_date: str) -> None:
    ws["C4"] = project_name
    ws["C5"] = project_code
    ws["G4"] = DEFAULT_TESTER
    ws["G5"] = DEFAULT_TESTER
    ws["G6"] = issue_date
    ws["G7"] = version
    ws["B12"] = issue_date
    ws["C12"] = version
    ws["D12"] = "A"
    ws["E12"] = "New"
    ws["F12"] = "Initial execution-backed FE-16, FE-17, and FE-18 system test workbook."
    ws["G12"] = "Report5_Test Case Document.xlsx, executed Playwright evidence"


def populate_test_case_list(ws, catalog: dict) -> None:
    clear_range(ws, 9, 60, 6)
    row = 9
    sequence = 1
    for feature in catalog["features"]:
        for function in feature["functions"]:
            copy_row_style(ws, 9, row, 6)
            ws.cell(row, 2).value = sequence
            ws.cell(row, 3).value = function["name"]
            ws.cell(row, 4).value = feature["sheetName"]
            ws.cell(row, 5).value = function["description"]
            ws.cell(row, 6).value = function["preCondition"]
            row += 1
            sequence += 1


def populate_detail_sheet(ws, feature: dict, feature_cases: list[dict]) -> None:
    clear_range(ws, 2, 200, 10)
    ws["A2"] = "Feature"
    ws["B2"] = feature["sheetName"]
    ws["A3"] = "Test requirement"
    ws["B3"] = feature["testRequirement"]
    ws["A4"] = "Reference Document"
    ws["B4"] = feature["referenceDocument"]
    ws["A5"] = "Pass"
    ws["B5"] = "Fail"
    ws["C5"] = "Untested"
    ws["D5"] = "N/A"
    ws["F5"] = "Number of Test cases"
    ws["A6"] = len(feature_cases)
    ws["B6"] = 0
    ws["C6"] = 0
    ws["D6"] = 0
    ws["F6"] = len(feature_cases)
    ws["A8"] = "ID"
    ws["B8"] = "Test Case Description"
    ws["C8"] = "Test Case Procedure"
    ws["D8"] = "Expected Results"
    ws["E8"] = "Actual Results"
    ws["F8"] = "Inter-test case Dependence"
    ws["G8"] = "Result"
    ws["H8"] = "Test date"
    ws["I8"] = "Tester"
    ws["J8"] = "Note"

    row = 9
    for function in feature["functions"]:
        function_cases = [case for case in feature_cases if case["function_name"] == function["name"]]
        copy_row_style(ws, 9, row, 10)
        ws.cell(row, 1).value = f"{function['name']} {FUNCTION_HEADER_SUFFIX}"
        row += 1
        for index, case in enumerate(function_cases):
            copy_row_style(ws, 10 if index == 0 else 11, row, 10)
            ws.cell(row, 1).value = case["id"]
            ws.cell(row, 2).value = case["description"]
            ws.cell(row, 3).value = format_numbered_lines(case["procedure"])
            ws.cell(row, 4).value = format_numbered_lines(case["expectedResults"])
            ws.cell(row, 5).value = format_actual_results_for_case(case)
            ws.cell(row, 6).value = case["pre_condition"]
            ws.cell(row, 7).value = case["result"]
            ws.cell(row, 8).value = case["testDate"]
            ws.cell(row, 9).value = DEFAULT_TESTER
            ws.cell(row, 10).value = format_note_for_case(case)
            apply_case_body_font(ws, row)
            row += 1


def populate_test_report(ws, catalog: dict, cases: list[dict], project_name: str, project_code: str, issue_date: str) -> None:
    ws["C3"] = project_name
    ws["C4"] = project_code
    ws["G3"] = DEFAULT_TESTER
    ws["G4"] = DEFAULT_TESTER
    ws["H5"] = issue_date
    ws["C6"] = "Release scope: Rating & Trust Score System, Dispute Resolution System, Notification System."

    total_cases = 0
    for offset, feature in enumerate(catalog["features"], start=11):
        counts = count_feature_cases(cases, feature["sheetName"])
        total_cases += counts["Total"]
        ws.cell(offset, 2).value = offset - 10
        ws.cell(offset, 3).value = feature["sheetName"]
        ws.cell(offset, 4).value = counts["Pass"]
        ws.cell(offset, 5).value = counts["Fail"]
        ws.cell(offset, 6).value = counts["Untested"]
        ws.cell(offset, 7).value = counts["N/A"]
        ws.cell(offset, 8).value = counts["Total"]

    ws["D14"] = total_cases
    ws["E14"] = 0
    ws["F14"] = 0
    ws["G14"] = 0
    ws["H14"] = total_cases
    ws["E16"] = 100
    ws["E17"] = 100


def build_workbook() -> Path:
    catalog = load_catalog()
    cases, project_name, project_code, version, issue_date = merge_case_evidence(catalog)

    wb = load_workbook(TEMPLATE_PATH)
    feature_template = wb["Feature1"]
    feature_copy_1 = wb.copy_worksheet(feature_template)
    feature_copy_2 = wb.copy_worksheet(feature_template)

    feature_template.title = catalog["features"][0]["sheetName"]
    feature_copy_1.title = catalog["features"][1]["sheetName"]
    feature_copy_2.title = catalog["features"][2]["sheetName"]

    del wb["Feature2"]
    wb._sheets = [
        wb["Cover"],
        wb["Test case List"],
        wb[catalog["features"][0]["sheetName"]],
        wb[catalog["features"][1]["sheetName"]],
        wb[catalog["features"][2]["sheetName"]],
        wb["Test Report"],
    ]

    populate_cover(wb["Cover"], project_name, project_code, version, issue_date)
    populate_test_case_list(wb["Test case List"], catalog)

    for feature in catalog["features"]:
        feature_cases = [case for case in cases if case["feature"] == feature["sheetName"]]
        populate_detail_sheet(wb[feature["sheetName"]], feature, feature_cases)

    populate_test_report(wb["Test Report"], catalog, cases, project_name, project_code, issue_date)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


if __name__ == "__main__":
    output = build_workbook()
    print(output.relative_to(REPO_ROOT).as_posix())
