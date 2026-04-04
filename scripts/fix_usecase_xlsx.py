from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
RAW_PATH = ROOT / "docs" / "use-cases-raw.txt"
OUT_XLSX_PATH = ROOT / "docs" / "use-cases-fixed.xlsx"
OUT_MAP_PATH = ROOT / "docs" / "use-cases-renumber-map.csv"


UC_CODE_RE = re.compile(r"^UC-(\d{1,3})$")


@dataclass(frozen=True)
class Row:
    original_index: int
    values: dict[str, str]

    @property
    def name(self) -> str:
        return (self.values.get("Tên UC") or "").strip()

    @property
    def code(self) -> str:
        return (self.values.get("Mã UC") or "").strip()


def _split_tsv_line(line: str) -> list[str]:
    return [cell.strip() for cell in line.rstrip("\n").split("\t")]


def _read_rows(raw_text: str) -> tuple[list[str], list[Row]]:
    lines = [ln for ln in raw_text.splitlines() if ln.strip()]
    if not lines:
        raise SystemExit("Raw file is empty.")

    header = _split_tsv_line(lines[0])
    rows: list[Row] = []
    for i, line in enumerate(lines[1:], start=1):
        cells = _split_tsv_line(line)
        if len(cells) < len(header):
            cells += [""] * (len(header) - len(cells))
        elif len(cells) > len(header):
            # Keep extra cells by extending header deterministically.
            for extra_i in range(len(header), len(cells)):
                header.append(f"Extra {extra_i - len(header) + 1}")
        values = {header[j]: cells[j] if j < len(cells) else "" for j in range(len(header))}
        rows.append(Row(original_index=i, values=values))
    return header, rows


def _is_uc_row(row: Row) -> bool:
    return bool(UC_CODE_RE.match(row.code))


def _normalize_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip().lower())


def _move_related_next_to_anchor(
    rows: list[Row],
    *,
    anchor_name: str,
    related_names: list[str],
) -> list[Row]:
    """
    Keep the original global order, but ensure `related_names` appear directly
    after `anchor_name` (in the given order), if they exist later in the list.
    """
    norm_anchor = _normalize_name(anchor_name)
    norm_related = [_normalize_name(nm) for nm in related_names]

    by_name: dict[str, list[Row]] = {}
    for r in rows:
        by_name.setdefault(_normalize_name(r.name), []).append(r)

    # Find anchor position (first match).
    anchor_idx = next((i for i, r in enumerate(rows) if _normalize_name(r.name) == norm_anchor), None)
    if anchor_idx is None:
        return rows

    # Collect related rows from *after* the anchor, removing them from their old positions.
    remaining: list[Row] = []
    pulled: list[Row] = []
    for i, r in enumerate(rows):
        if i <= anchor_idx:
            remaining.append(r)
            continue
        n = _normalize_name(r.name)
        if n in norm_related:
            pulled.append(r)
        else:
            remaining.append(r)

    # Rebuild: everything up to anchor, then pulled in requested order, then the rest.
    head = remaining[: anchor_idx + 1]
    tail = remaining[anchor_idx + 1 :]

    pulled_by_norm: dict[str, list[Row]] = {}
    for r in pulled:
        pulled_by_norm.setdefault(_normalize_name(r.name), []).append(r)

    ordered_pulled: list[Row] = []
    for rn in norm_related:
        ordered_pulled.extend(pulled_by_norm.get(rn, []))

    return head + ordered_pulled + tail


def _format_uc_code(n: int) -> str:
    if n < 100:
        return f"UC-{n:02d}"
    return f"UC-{n}"


def _write_xlsx(header: list[str], rows: list[Row], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "UseCases"

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)

    ws.append(header)
    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment

    for row in rows:
        ws.append([row.values.get(col, "") for col in header])

    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(1, len(header) + 1):
            ws.cell(row=row_idx, column=col_idx).alignment = cell_alignment

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}1"

    # Column widths: heuristic based on content (capped).
    for col_idx, col_name in enumerate(header, start=1):
        max_len = len(col_name)
        for row_idx in range(2, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    raw_text = RAW_PATH.read_text(encoding="utf-8")
    header, all_rows = _read_rows(raw_text)

    uc_rows = [r for r in all_rows if _is_uc_row(r)]
    if not uc_rows:
        raise SystemExit("No UC rows found. Expected 'Mã UC' like UC-01.")

    # Minimal, safe reordering: keep the original flow but pull related UCs together.
    fixed_order = list(uc_rows)
    fixed_order = _move_related_next_to_anchor(
        fixed_order,
        anchor_name="View Profile",
        related_names=["Edit Profile", "Delete Account"],
    )

    # Renumber sequentially.
    renumber_map: list[tuple[str, str, str]] = []  # old_code, new_code, name
    for i, row in enumerate(fixed_order, start=1):
        old_code = row.values.get("Mã UC", "")
        old_no = row.values.get("F", "")
        new_code = _format_uc_code(i)
        row.values["F"] = str(i)
        row.values["Mã UC"] = new_code
        renumber_map.append((old_code, new_code, row.name or old_no))

    _write_xlsx(header, fixed_order, OUT_XLSX_PATH)

    OUT_MAP_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_MAP_PATH.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["old_uc", "new_uc", "name"])
        w.writerows(renumber_map)

    print(f"Wrote: {OUT_XLSX_PATH}")
    print(f"Wrote: {OUT_MAP_PATH}")


if __name__ == "__main__":
    main()
