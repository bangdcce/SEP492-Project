from __future__ import annotations

from copy import copy
from pathlib import Path
import shutil
import sys

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter


WORKBOOK_PATH = Path(r"D:\GradProject\SEP492-Project\docs\Unit Test.xlsx")
BACKUP_PATH = Path(r"D:\GradProject\SEP492-Project\docs\Unit Test.layout-sync.backup.xlsx")
FALLBACK_PATH = Path(r"D:\GradProject\SEP492-Project\docs\Unit Test.layout-sync.xlsx")

TOP_MERGES = [
    "A2:B2",
    "C2:E2",
    "F2:K2",
    "L2:T2",
    "A3:B3",
    "C3:E3",
    "F3:K3",
    "L3:T3",
    "A4:B4",
    "C4:D4",
    "F4:K4",
    "L4:T4",
    "A5:B5",
    "C5:T5",
    "A6:B6",
    "C6:E6",
    "F6:K6",
    "L6:N6",
    "O6:T6",
    "A7:B7",
    "C7:E7",
    "F7:K7",
    "O7:T7",
]


def classify_sheet(ws) -> str:
    if ws["F2"].value == "Function Name":
        return "standard"
    if ws["E2"].value == "Function Name":
        return "shift_e"
    if ws["D2"].value == "Function Name":
        return "shift_d"
    if ws["D1"].value == "Function Name":
        return "compact"
    return "unknown"


def is_case_mark(value) -> bool:
    return isinstance(value, str) and value in {"O", "P", "F", "U", "x", "X"}


def translate_formula(value, src_coord: str, dest_coord: str):
    if isinstance(value, str) and value.startswith("="):
        try:
            return Translator(value, origin=src_coord).translate_formula(dest_coord)
        except Exception:
            return value
    return value


def map_coord(sheet_type: str, row: int, col: int):
    new_row = row
    new_col = col

    if sheet_type == "shift_e":
        if row in {2, 3, 4, 6, 7} and col >= 5:
            new_col = col + 1
        elif row >= 9:
            if col == 3:
                new_col = 4
            elif col >= 5:
                new_col = col + 1
        return new_row, new_col

    if sheet_type == "shift_d":
        if row in {2, 3, 4, 6, 7} and col >= 4:
            new_col = col + 2
        elif row >= 9:
            if col == 3:
                new_col = 4
            elif col >= 4:
                new_col = col + 2
        return new_row, new_col

    if sheet_type == "compact":
        new_row = row + 1
        if row in {1, 2, 3} and col >= 4:
            new_col = col + 2
        elif row in {5, 6} and col >= 4:
            new_col = col + 2
        elif row >= 8:
            if col == 3:
                new_col = 4
            elif col >= 4:
                new_col = col + 2
        return new_row, new_col

    return new_row, new_col


def collect_sheet_payload(ws, sheet_type: str):
    payload = []
    row_heights = {}
    max_case_col = 0

    for row_idx in range(1, ws.max_row + 1):
        row_heights[row_idx] = ws.row_dimensions[row_idx].height
        for cell in ws[row_idx]:
            if (
                cell.value in (None, "")
                and cell.comment is None
                and cell.hyperlink is None
                and not cell.has_style
            ):
                continue

            dest_row, dest_col = map_coord(sheet_type, cell.row, cell.column)
            if is_case_mark(cell.value):
                max_case_col = max(max_case_col, dest_col)
            payload.append(
                {
                    "src": cell.coordinate,
                    "dest_row": dest_row,
                    "dest_col": dest_col,
                    "value": cell.value,
                    "style": copy(cell._style),
                    "number_format": cell.number_format,
                    "font": copy(cell.font),
                    "fill": copy(cell.fill),
                    "border": copy(cell.border),
                    "alignment": copy(cell.alignment),
                    "protection": copy(cell.protection),
                    "hyperlink": copy(cell.hyperlink) if cell.hyperlink else None,
                    "comment": copy(cell.comment) if cell.comment else None,
                }
            )
    return payload, row_heights, max_case_col


def apply_template_layout(ws, template):
    # widths A:T
    for col_idx in range(1, 21):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = template.column_dimensions[letter].width

    # row heights for top/template rows
    for row_idx in range(2, 11):
        ws.row_dimensions[row_idx].height = template.row_dimensions[row_idx].height

    # top-cell styles A:T on rows 2:10
    for row_idx in range(2, 11):
        for col_idx in range(1, 21):
            ws.cell(row_idx, col_idx)._style = copy(template.cell(row_idx, col_idx)._style)

    for merge in TOP_MERGES:
        ws.merge_cells(merge)


def clear_sheet(ws):
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell._style = copy(ws.parent["EP-01 Register Account"]["A1"]._style)
            cell.hyperlink = None
            cell.comment = None


def locate_row(ws, label: str):
    for row_idx in range(1, ws.max_row + 1):
        if ws.cell(row_idx, 1).value == label or ws.cell(row_idx, 2).value == label:
            return row_idx
    return None


def style_summary_rows(ws, template):
    result_row = locate_row(ws, "Result")
    passed_row = locate_row(ws, "Passed/Failed")
    executed_row = locate_row(ws, "Executed Date")
    defect_row = locate_row(ws, "Defect ID")
    rows = [result_row, passed_row, executed_row, defect_row]
    template_rows = [66, 67, 68, 69]

    for target_row, template_row in zip(rows, template_rows):
        if not target_row:
            continue
        for col_idx in range(1, 21):
            ws.cell(target_row, col_idx)._style = copy(template.cell(template_row, col_idx)._style)
        ws.row_dimensions[target_row].height = template.row_dimensions[template_row].height
        ws.merge_cells(f"B{target_row}:D{target_row}")


def write_payload(ws, payload):
    for item in payload:
        cell = ws.cell(item["dest_row"], item["dest_col"])
        cell.value = translate_formula(item["value"], item["src"], cell.coordinate)
        cell._style = copy(item["style"])
        cell.number_format = item["number_format"]
        cell.font = copy(item["font"])
        cell.fill = copy(item["fill"])
        cell.border = copy(item["border"])
        cell.alignment = copy(item["alignment"])
        cell.protection = copy(item["protection"])
        if item["hyperlink"]:
            cell._hyperlink = copy(item["hyperlink"])
        if item["comment"]:
            cell.comment = copy(item["comment"])


def set_row_heights(ws, row_heights, sheet_type: str):
    for src_row, height in row_heights.items():
        if height is None:
            continue
        dest_row, _ = map_coord(sheet_type, src_row, 1)
        # top rows and summary rows get template heights later
        if dest_row < 11:
            continue
        ws.row_dimensions[dest_row].height = height


def ensure_utc_headers(ws, case_count: int):
    if case_count <= 0:
        return
    for idx in range(case_count):
        ws.cell(9, 6 + idx).value = f"UTCID{idx + 1:02d}"


def normalize_sheet(wb, sheet_name: str, template):
    old_ws = wb[sheet_name]
    sheet_type = classify_sheet(old_ws)
    if sheet_type in {"standard", "unknown"}:
        return False, sheet_type

    payload, row_heights, max_case_col = collect_sheet_payload(old_ws, sheet_type)
    sheet_index = wb.index(old_ws)
    temp_title = f"__tmp__{sheet_name[:20]}"
    if temp_title in wb.sheetnames:
        wb.remove(wb[temp_title])
    new_ws = wb.create_sheet(temp_title, sheet_index)
    clear_sheet(new_ws)
    apply_template_layout(new_ws, template)
    write_payload(new_ws, payload)
    set_row_heights(new_ws, row_heights, sheet_type)

    # Standardize UTC header row for nonstandard sheets.
    case_count = max(0, max_case_col - 5)
    ensure_utc_headers(new_ws, case_count)

    # Standardize key header labels.
    new_ws["A2"] = "Function Code"
    new_ws["F2"] = "Function Name"
    new_ws["A3"] = "Created By"
    new_ws["F3"] = "Executed By"
    new_ws["A4"] = "Lines  of code"
    new_ws["F4"] = "Lack of test cases"
    new_ws["A5"] = "Test requirement"
    new_ws["A6"] = "Passed"
    new_ws["C6"] = "Failed"
    new_ws["F6"] = "Untested"
    new_ws["L6"] = "N/A/B"
    new_ws["O6"] = "Total Test Cases"
    new_ws["A10"] = "Condition"
    new_ws["B10"] = "Precondition"

    style_summary_rows(new_ws, template)

    # keep same title/order
    wb.remove(old_ws)
    new_ws.title = sheet_name
    return True, sheet_type


def main():
    shutil.copy2(WORKBOOK_PATH, BACKUP_PATH)
    wb = load_workbook(WORKBOOK_PATH)
    template = wb["EP-01 Register Account"]

    changed = []
    skipped = []
    for sheet_name in list(wb.sheetnames):
        if not sheet_name.startswith("EP-"):
            continue
        did_change, sheet_type = normalize_sheet(wb, sheet_name, template)
        if did_change:
            changed.append((sheet_name, sheet_type))
        else:
            skipped.append((sheet_name, sheet_type))

    try:
        wb.save(WORKBOOK_PATH)
        saved_path = WORKBOOK_PATH
    except PermissionError:
        wb.save(FALLBACK_PATH)
        saved_path = FALLBACK_PATH

    check = load_workbook(saved_path)
    print("saved", saved_path)
    print("backup", BACKUP_PATH)
    print("changed", len(changed))
    print("skipped", len(skipped))
    print("sample_sheets", ["EP-03 Create Request", "EP-08 Edit Request", "EP-20 Get Overview"])
    for name in ["EP-03 Create Request", "EP-08 Edit Request", "EP-20 Get Overview", "EP-174 Verify Email"]:
        if name not in check.sheetnames:
            continue
        ws = check[name]
        print(
            name,
            ws["A2"].value,
            ws["C2"].value,
            ws["F2"].value,
            ws["L2"].value,
            ws["F9"].value,
            ws["A10"].value,
            ws["B10"].value,
        )


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
